# ════════════════════════════════════════════════════════════════════════════
# TAM DASHBOARD PAGE
# Paste this as elif tab == X in Home.py
# Add ("👤", "TAM") to TABS list
# ════════════════════════════════════════════════════════════════════════════

elif tab == 10:
    import openpyxl
    import pandas as pd
    from pathlib import Path
    from utils.charts import COLORS, bar_chart, donut_chart

    st.markdown("""
<div style='display:flex;align-items:center;gap:12px;margin-bottom:4px'>
  <div style='font-size:20px;font-weight:700;color:#e6edf3'>👤 TAM / TSE Account Dashboard</div>
  <div style='font-size:11px;color:#8b949e;margin-top:2px'>
    Source: accounts.xlsx · Select a TSE to view their portfolio
  </div>
</div>""", unsafe_allow_html=True)

    # ── Load accounts.xlsx ────────────────────────────────────────────────────
    ACCOUNTS_FILE = Path("data/accounts.xlsx")
    if not ACCOUNTS_FILE.exists():
        st.error(
            "accounts.xlsx not found at `data/accounts.xlsx`. "
            "Copy the file from Downloads:\n"
            "```bash\nmkdir -p data && cp ~/Downloads/accounts.xlsx data/\n```"
        )
        st.stop()

    @st.cache_data(ttl=3600, show_spinner=False)
    def load_accounts():
        wb  = openpyxl.load_workbook("data/accounts.xlsx", read_only=True, data_only=True)

        # ── Up to date sheet ──────────────────────────────────────────────
        ws  = wb["Up to date"]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip() if h else f"col_{i}" for i,h in enumerate(rows[0])]
        data = []
        for row in rows[1:]:
            if any(v is not None for v in row):
                data.append(dict(zip(headers, row)))
        df_main = pd.DataFrame(data)

        # ── Account Contact List ──────────────────────────────────────────
        ws2 = wb["Account Contact List"]
        rows2 = list(ws2.iter_rows(values_only=True))
        h2 = [str(h).strip() if h else f"col_{i}" for i,h in enumerate(rows2[0])]
        d2 = [dict(zip(h2, r)) for r in rows2[1:] if any(v is not None for v in r)]
        df_contacts = pd.DataFrame(d2)

        # ── Ongoing Projects ──────────────────────────────────────────────
        ws3 = wb["Ongoing Projects"]
        rows3 = list(ws3.iter_rows(values_only=True))
        h3 = [str(h).strip() if h else f"col_{i}" for i,h in enumerate(rows3[1])]
        d3 = [dict(zip(h3, r)) for r in rows3[2:] if any(v is not None for v in r)]
        df_projects = pd.DataFrame(d3)

        # ── Incoming ──────────────────────────────────────────────────────
        ws4 = wb["Incoming"]
        rows4 = list(ws4.iter_rows(values_only=True))
        h4 = [str(h).strip() if h else f"col_{i}" for i,h in enumerate(rows4[1])]
        d4 = [dict(zip(h4, r)) for r in rows4[2:] if any(v is not None for v in r)]
        df_incoming = pd.DataFrame(d4)

        wb.close()
        return df_main, df_contacts, df_projects, df_incoming

    with st.spinner("Loading account data..."):
        df_main, df_contacts, df_projects, df_incoming = load_accounts()

    # ── Normalize key columns ─────────────────────────────────────────────────
    def col(df, *candidates):
        """Find first matching column name."""
        for c in candidates:
            if c in df.columns:
                return c
            matches = [x for x in df.columns if c.lower() in str(x).lower()]
            if matches:
                return matches[0]
        return None

    tse_col     = col(df_main, "Assigned TSE", "TSE")
    csm_col     = col(df_main, "Customer Success Manager", "CSM")
    cust_col    = col(df_main, "Customer name", "Customer")
    arr_col     = col(df_main, "ARR")
    renewal_col = col(df_main, "Renewal Date")
    stage_col   = col(df_main, "Lifecycle Stage")
    core_col    = col(df_main, "Core Status")
    assign_col  = col(df_main, "Assign Status", "Assign purchased")
    assist_col  = col(df_main, "Assist Status", "Assist purchased")
    expand_col  = col(df_main, "Expand Status", "Expand purchased")
    elevate_col = col(df_main, "Elevate Status", "Elevate Customers", "Elevate purchased")
    resolve_col = col(df_main, "Resolve Status", "Resolve purchased")
    crm_col     = col(df_main, "CRM Importer", "SOR")
    sso_col     = col(df_main, "SSO Provider")
    wb_col      = col(df_main, "Write backs")
    notes_col   = col(df_main, "Notes")
    url_col     = col(df_main, "Customer URLs")

    # ── TSE selector ──────────────────────────────────────────────────────────
    all_tses = sorted(df_main[tse_col].dropna().unique().tolist()) if tse_col else []
    col_tse, col_csm, _ = st.columns([1, 1, 2])
    selected_tse = col_tse.selectbox("Select TSE", ["All"] + all_tses, key="tam_tse")
    all_csms = sorted(df_main[csm_col].dropna().unique().tolist()) if csm_col else []
    selected_csm = col_csm.selectbox("Filter by CSM", ["All"] + all_csms, key="tam_csm")

    # Filter
    df_filtered = df_main.copy()
    if selected_tse != "All" and tse_col:
        df_filtered = df_filtered[df_filtered[tse_col].astype(str).str.strip() == selected_tse]
    if selected_csm != "All" and csm_col:
        df_filtered = df_filtered[df_filtered[csm_col].astype(str).str.strip() == selected_csm]
    df_filtered = df_filtered[df_filtered[cust_col].notna()] if cust_col else df_filtered

    st.divider()

    # ── Portfolio KPIs ────────────────────────────────────────────────────────
    total_accounts = len(df_filtered)
    total_arr = 0
    if arr_col:
        arr_vals = pd.to_numeric(df_filtered[arr_col], errors='coerce').dropna()
        total_arr = arr_vals.sum()

    live_count = 0
    if stage_col:
        live_count = df_filtered[stage_col].astype(str).str.lower().str.contains("adopt|live|impl").sum()

    # Renewals in next 90 days
    renewal_soon = 0
    if renewal_col:
        today = pd.Timestamp.now()
        def parse_renewal(v):
            if isinstance(v, datetime_type := __import__('datetime').datetime):
                return pd.Timestamp(v)
            try:
                return pd.to_datetime(v)
            except:
                return None
        import datetime
        renewals = df_filtered[renewal_col].apply(parse_renewal)
        renewal_soon = ((renewals >= today) & (renewals <= today + pd.Timedelta(days=90))).sum()

    c1,c2,c3,c4,c5 = st.columns(5)
    c1.metric("Total accounts", total_accounts)
    c2.metric("Total ARR", f"${total_arr/1000:.0f}K" if total_arr else "—")
    c3.metric("Live / Adoption", live_count)
    c4.metric("Renewals in 90d", renewal_soon, "⚠️" if renewal_soon > 0 else "")
    # Projects for this TSE
    proj_count = 0
    if "Project Owner" in df_projects.columns and selected_tse != "All":
        proj_count = df_projects[df_projects["Project Owner"].astype(str).str.lower().str.contains(
            selected_tse.lower().split()[0] if selected_tse else "", na=False)].shape[0]
    c5.metric("Active projects", proj_count)

    st.divider()

    # ── Account cards ─────────────────────────────────────────────────────────
    st.subheader(f"📋 Account portfolio — {selected_tse if selected_tse != 'All' else 'All TSEs'}")

    STATUS_ICON = {
        "live":         ("✅", "#1a3a2a", "#1D9E75"),
        "implementing": ("🔧", "#2a2a10", "#BA7517"),
        "entitled":     ("📋", "#1a1a3a", "#534AB7"),
        "not":          ("⬜", "#1e1e1e", "#444"),
        "churned":      ("❌", "#3a1a1a", "#E24B4A"),
        "scoping":      ("🔍", "#1a1a3a", "#185FA5"),
        "other":        ("⚪", "#161b22", "#30363d"),
    }

    def status_class(val):
        if not val or str(val).strip() in ("--","None","nan",""):
            return "not"
        v = str(val).lower()
        if "churn" in v:  return "churned"
        if "live" in v:   return "live"
        if "impl" in v:   return "implementing"
        if "entitl" in v: return "entitled"
        if "scop" in v:   return "scoping"
        return "other"

    def module_badge(label, val):
        sc = status_class(val)
        icon, bg, border = STATUS_ICON[sc]
        disp = str(val).replace("--","—") if val and str(val) not in ("nan","None","--") else "—"
        return (f"<div style='background:{bg};border:0.5px solid {border};"
                f"border-radius:6px;padding:4px 8px;font-size:10px;margin:2px'>"
                f"<span style='color:{border}'>{icon}</span> "
                f"<span style='color:#8b949e'>{label}</span> "
                f"<span style='color:#e6edf3;font-weight:500'>{disp}</span></div>")

    for _, row in df_filtered.iterrows():
        cust_name = str(row.get(cust_col,"")).strip() if cust_col else "—"
        if not cust_name or cust_name in ("nan","None"):
            continue

        stage  = str(row.get(stage_col,"")).strip()  if stage_col  else ""
        arr    = row.get(arr_col)                     if arr_col    else None
        renew  = row.get(renewal_col)                 if renewal_col else None
        tse    = str(row.get(tse_col,"")).strip()     if tse_col    else "—"
        csm    = str(row.get(csm_col,"")).strip()     if csm_col    else "—"
        notes  = str(row.get(notes_col,"")).strip()   if notes_col  else ""
        url    = str(row.get(url_col,"")).strip()     if url_col    else ""
        crm    = str(row.get(crm_col,"")).strip()     if crm_col    else "—"
        sso    = str(row.get(sso_col,"")).strip()     if sso_col    else "—"
        wb_val = str(row.get(wb_col,"")).strip()      if wb_col     else "—"

        # Renewal date formatting
        renew_str = "—"
        renew_flag = ""
        if renew and str(renew) not in ("nan","None",""):
            try:
                import datetime
                if isinstance(renew, datetime.datetime):
                    rdt = pd.Timestamp(renew)
                else:
                    rdt = pd.to_datetime(renew)
                renew_str = rdt.strftime("%b %Y")
                days_to = (rdt - pd.Timestamp.now()).days
                if days_to < 0:
                    renew_flag = "🔴 Overdue"
                elif days_to <= 90:
                    renew_flag = f"🟡 {days_to}d away"
                else:
                    renew_flag = f"🟢 {days_to}d away"
            except:
                renew_str = str(renew)[:10]

        # ARR formatting
        arr_str = f"${float(arr)/1000:.0f}K" if arr and str(arr) not in ("nan","None") else "—"

        # Stage color
        stage_lower = stage.lower()
        if "adopt" in stage_lower or "live" in stage_lower:
            stage_color = "#1D9E75"
        elif "impl" in stage_lower:
            stage_color = "#BA7517"
        elif "churn" in stage_lower:
            stage_color = "#E24B4A"
        else:
            stage_color = "#534AB7"

        # Module statuses
        modules = {
            "Core":    row.get(core_col)    if core_col    else None,
            "Assign":  row.get(assign_col)  if assign_col  else None,
            "Assist":  row.get(assist_col)  if assist_col  else None,
            "Expand":  row.get(expand_col)  if expand_col  else None,
            "Elevate": row.get(elevate_col) if elevate_col else None,
            "Resolve": row.get(resolve_col) if resolve_col else None,
        }

        # Notes truncation
        notes_disp = (notes[:180] + "…") if len(notes) > 180 else notes
        notes_disp = notes_disp if notes_disp not in ("nan","None","") else ""

        url_link = f"<a href='{url}' target='_blank' style='color:#185FA5;font-size:10px'>🔗 Open SL portal</a>" if url and url not in ("nan","None") else ""

        badges = "".join(module_badge(k,v) for k,v in modules.items())

        with st.expander(
            f"**{cust_name}**  ·  "
            f"<span style='color:{stage_color}'>{stage}</span>  ·  "
            f"{arr_str}  ·  Renewal: {renew_str} {renew_flag}  ·  TSE: {tse}",
            expanded=False
        ):
            col_l, col_r = st.columns([3, 2])
            with col_l:
                st.markdown(f"""
<div style='background:#161b22;border-radius:8px;padding:12px 16px;margin-bottom:8px'>
  <div style='display:flex;gap:24px;flex-wrap:wrap;margin-bottom:8px'>
    <div><span style='font-size:10px;color:#8b949e'>CSM</span><br><span style='color:#e6edf3;font-size:12px'>{csm}</span></div>
    <div><span style='font-size:10px;color:#8b949e'>TSE</span><br><span style='color:#e6edf3;font-size:12px'>{tse}</span></div>
    <div><span style='font-size:10px;color:#8b949e'>ARR</span><br><span style='color:#1D9E75;font-size:12px;font-weight:600'>{arr_str}</span></div>
    <div><span style='font-size:10px;color:#8b949e'>Renewal</span><br><span style='color:#e6edf3;font-size:12px'>{renew_str}</span> <span style='font-size:10px'>{renew_flag}</span></div>
    <div><span style='font-size:10px;color:#8b949e'>CRM</span><br><span style='color:#e6edf3;font-size:12px'>{crm}</span></div>
    <div><span style='font-size:10px;color:#8b949e'>SSO</span><br><span style='color:#e6edf3;font-size:12px'>{sso}</span></div>
    <div><span style='font-size:10px;color:#8b949e'>Writeback</span><br><span style='color:#e6edf3;font-size:12px'>{wb_val}</span></div>
  </div>
  {url_link}
  {"<div style='margin-top:8px;font-size:11px;color:#8b949e;line-height:1.5'>" + notes_disp + "</div>" if notes_disp else ""}
</div>""", unsafe_allow_html=True)

            with col_r:
                st.markdown("**Module status**")
                st.markdown(f"<div style='display:flex;flex-wrap:wrap;gap:2px'>{badges}</div>",
                            unsafe_allow_html=True)

    st.divider()

    # ── Ongoing Projects ──────────────────────────────────────────────────────
    st.subheader("🔧 Ongoing projects")
    if not df_projects.empty:
        proj_col_owner = col(df_projects, "Project Owner")
        proj_col_cust  = col(df_projects, "Customer")
        proj_col_prog  = col(df_projects, "Progress")
        proj_col_task  = col(df_projects, "Task/Activity", "Task")
        proj_col_status= col(df_projects, "Current Status")
        proj_col_next  = col(df_projects, "Next Steps")
        proj_col_start = col(df_projects, "Start Date")
        proj_col_age   = col(df_projects, "Project Age")

        df_proj_show = df_projects.copy()
        if selected_tse != "All" and proj_col_owner:
            df_proj_show = df_proj_show[
                df_proj_show[proj_col_owner].astype(str).str.lower().str.contains(
                    selected_tse.lower().split()[0], na=False)]

        STATUS_COLORS = {
            "in progress": "#1D9E75",
            "not started": "#BA7517",
            "scoping":     "#185FA5",
            "done":        "#534AB7",
            "blocked":     "#E24B4A",
        }

        for _, proj in df_proj_show.iterrows():
            owner  = str(proj.get(proj_col_owner,"")).strip()  if proj_col_owner  else "—"
            pcust  = str(proj.get(proj_col_cust,"")).strip()   if proj_col_cust   else "—"
            prog   = str(proj.get(proj_col_prog,"")).strip()   if proj_col_prog   else "—"
            task   = str(proj.get(proj_col_task,"")).strip()   if proj_col_task   else "—"
            status = str(proj.get(proj_col_status,"")).strip() if proj_col_status else ""
            nexts  = str(proj.get(proj_col_next,"")).strip()   if proj_col_next   else ""

            if task in ("nan","None","") and pcust in ("nan","None",""):
                continue

            prog_c = STATUS_COLORS.get(prog.lower(), "#8b949e")
            st.markdown(f"""
<div style='background:#161b22;border:0.5px solid #30363d;border-radius:8px;
     padding:12px 16px;margin-bottom:6px;border-left:3px solid {prog_c}'>
  <div style='display:flex;justify-content:space-between;align-items:flex-start'>
    <div>
      <span style='font-size:13px;font-weight:500;color:#e6edf3'>{task if task not in ("nan","None") else "—"}</span>
      <span style='font-size:11px;color:#8b949e;margin-left:10px'>{pcust if pcust not in ("nan","None") else ""}</span>
    </div>
    <div style='display:flex;gap:8px;align-items:center'>
      <span style='font-size:10px;color:{prog_c};font-weight:600'>{prog}</span>
      <span style='font-size:10px;color:#8b949e'>{owner}</span>
    </div>
  </div>
  {"<div style='font-size:11px;color:#8b949e;margin-top:6px'>📌 " + status + "</div>" if status and status not in ("nan","None") else ""}
  {"<div style='font-size:11px;color:#185FA5;margin-top:4px'>→ " + nexts + "</div>" if nexts and nexts not in ("nan","None") else ""}
</div>""", unsafe_allow_html=True)
    else:
        st.info("No ongoing projects data available.", icon="ℹ️")

    st.divider()

    # ── Incoming projects ─────────────────────────────────────────────────────
    st.subheader("📥 Incoming / pipeline")
    if not df_incoming.empty:
        inc_cust = col(df_incoming, "Customer")
        inc_task = col(df_incoming, "Task/Activity", "Task")
        inc_prog = col(df_incoming, "Progress")
        inc_pri  = col(df_incoming, "Priority")
        inc_own  = col(df_incoming, "Potential Project Owner")

        for _, inc in df_incoming.iterrows():
            icust = str(inc.get(inc_cust,"")).strip() if inc_cust else "—"
            itask = str(inc.get(inc_task,"")).strip() if inc_task else "—"
            iprog = str(inc.get(inc_prog,"")).strip() if inc_prog else "—"
            ipri  = str(inc.get(inc_pri,"")).strip()  if inc_pri  else "—"
            iown  = str(inc.get(inc_own,"")).strip()  if inc_own  else "—"
            if icust in ("nan","None","") and itask in ("nan","None",""):
                continue
            pri_c = "#E24B4A" if ipri.lower()=="high" else "#BA7517" if ipri.lower()=="medium" else "#8b949e"
            st.markdown(f"""
<div style='background:#161b22;border:0.5px solid #30363d;border-radius:8px;
     padding:10px 16px;margin-bottom:6px'>
  <div style='display:flex;justify-content:space-between'>
    <span style='font-size:12px;font-weight:500;color:#e6edf3'>
      {icust if icust not in ("nan","None") else "—"} — {itask if itask not in ("nan","None") else "—"}
    </span>
    <span style='font-size:10px;color:{pri_c};font-weight:600'>{ipri} priority</span>
  </div>
  <div style='font-size:10px;color:#8b949e;margin-top:4px'>{iprog} · Owner: {iown}</div>
</div>""", unsafe_allow_html=True)
    else:
        st.info("No incoming projects data.", icon="ℹ️")

    st.divider()

    # ── Portfolio summary charts ──────────────────────────────────────────────
    st.subheader("📊 Portfolio overview")
    col_ch1, col_ch2, col_ch3 = st.columns(3)

    # Stage breakdown
    if stage_col and not df_filtered.empty:
        stage_counts = df_filtered[stage_col].value_counts().reset_index()
        stage_counts.columns = ["stage","count"]
        stage_counts = stage_counts[stage_counts["stage"].notna()]
        with col_ch1:
            st.markdown("**Lifecycle stage**")
            st.plotly_chart(donut_chart(
                labels=stage_counts["stage"].tolist(),
                values=stage_counts["count"].tolist(),
                colors=[COLORS["teal"],COLORS["blue"],COLORS["amber"],
                        COLORS["red"],COLORS["gray"],COLORS["purple"]],
                height=220,
                center_text=f"{len(df_filtered)} accounts"
            ), use_container_width=True)

    # Module adoption
    module_live = {}
    for mod_name, mod_col in [
        ("Core", core_col), ("Assign", assign_col), ("Assist", assist_col),
        ("Expand", expand_col), ("Elevate", elevate_col), ("Resolve", resolve_col)
    ]:
        if mod_col and mod_col in df_filtered.columns:
            live = df_filtered[mod_col].astype(str).str.lower().str.contains("live|active").sum()
            module_live[mod_name] = int(live)

    if module_live:
        mod_df = pd.DataFrame(list(module_live.items()), columns=["module","live_count"])
        with col_ch2:
            st.markdown("**Modules live across portfolio**")
            st.plotly_chart(bar_chart(
                mod_df, x="module",
                y_cols=[{"col":"live_count","name":"Live accounts",
                          "color":[COLORS["teal"],COLORS["blue"],COLORS["purple"],
                                   COLORS["amber"],COLORS["green"],COLORS["red"]],
                          "colors":[COLORS["teal"],COLORS["blue"],COLORS["purple"],
                                    COLORS["amber"],COLORS["green"],COLORS["red"]]}],
                height=220
            ), use_container_width=True)

    # ARR by account (top 10)
    if arr_col and not df_filtered.empty:
        arr_df = df_filtered[[cust_col, arr_col]].copy()
        arr_df.columns = ["customer","arr"]
        arr_df["arr"] = pd.to_numeric(arr_df["arr"], errors="coerce")
        arr_df = arr_df.dropna().sort_values("arr", ascending=False).head(10)
        arr_df["arr_k"] = (arr_df["arr"]/1000).round(0)
        arr_df["customer"] = arr_df["customer"].str[:20]
        with col_ch3:
            st.markdown("**ARR by account (top 10, $K)**")
            st.plotly_chart(bar_chart(
                arr_df.sort_values("arr_k"), x="customer",
                y_cols=[{"col":"arr_k","name":"ARR ($K)","color":COLORS["teal"]}],
                height=220, horizontal=True
            ), use_container_width=True)

    st.caption(f"Source: data/accounts.xlsx · {len(df_filtered)} accounts shown · "
               f"TSE filter: {selected_tse} · CSM filter: {selected_csm}")
