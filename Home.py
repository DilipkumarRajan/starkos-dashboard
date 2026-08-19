"""
Home.py — SupportLogic Account Performance Dashboard
Single-page app with top navigation bar. All pages rendered inline.
Run: streamlit run Home.py
"""

import streamlit as st
import pandas as pd
import os
from pathlib import Path

st.set_page_config(
    page_title="STARKos — Support Intelligence",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Remove sidebar & default padding ─────────────────────────────────────────
st.markdown("""
<style>
[data-testid="collapsedControl"] { display:none }
[data-testid="stSidebar"] { display:none }
section.main > div { padding-top: 0.5rem !important; }
.block-container { padding-top: 0.5rem !important; padding-bottom: 1rem !important; }

/* Top navbar */
.sl-navbar {
  display: flex;
  align-items: center;
  gap: 0;
  background: #0A1931;
  border-radius: 10px;
  padding: 8px 16px;
  margin-bottom: 18px;
  flex-wrap: wrap;
  gap: 4px;
}
.sl-brand {
  font-size: 15px;
  font-weight: 600;
  color: #0F6E56;
  margin-right: 16px;
  white-space: nowrap;
}
.sl-tab {
  padding: 6px 14px;
  border-radius: 20px;
  border: none;
  font-size: 12px;
  font-weight: 500;
  cursor: pointer;
  white-space: nowrap;
  transition: all 0.15s;
  background: transparent;
  color: #8b949e;
  font-family: var(--font-sans);
}
.sl-tab:hover { background: rgba(255,255,255,0.08); color: #e6edf3; }
.sl-tab-active { background: #185FA5 !important; color: #ffffff !important; }
.sl-tab-0  { }
.sl-tab-1  { }
.sl-tab-2  { }
.sl-tab-3  { }
.sl-tab-4  { }
.sl-tab-5  { }
.sl-tab-6  { }
.sl-tab-7  { }

/* Metric cards */
div[data-testid="metric-container"] {
  background: #161b22;
  border-radius: 8px;
  padding: 12px;
  border: 0.5px solid #30363d;
}
/* Hide streamlit header */
header[data-testid="stHeader"] { display:none }
</style>
""", unsafe_allow_html=True)

# ── Customer selector & tab state ─────────────────────────────────────────────
from utils.customers import CUSTOMERS

if "active_tab" not in st.session_state:
    st.session_state.active_tab = 0
if "selected_customer" not in st.session_state:
    st.session_state.selected_customer = list(CUSTOMERS.keys())[0]

TABS = [
    ("📋", "Snap"),
    ("🧠", "Brief"),
    ("🛡️", "Esc"),
    ("😊", "Sent"),
    ("🔀", "Route"),
    ("🔬", "Feat"),
    ("📈", "ROI"),
    ("📱", "Pendo"),
    ("⚙️", "SQL"),
    ("🔍", "Logs"),
    ("👤", "TAM"),
]

# ── STARKos brand + Top navbar ───────────────────────────────────────────────
st.markdown("""
<div style='background:#0A1931;border-radius:10px;padding:10px 20px;margin-bottom:8px;
     display:flex;align-items:center;gap:12px;border:0.5px solid #1e2d4a'>
  <div style='font-size:24px;font-weight:800;letter-spacing:1px;
       font-family:Google Sans,Arial,sans-serif;line-height:1'>
    <span style='color:#E24B4A'>S</span><span style='color:#BA7517'>T</span><span style='color:#1D9E75'>A</span><span style='color:#185FA5'>R</span><span style='color:#534AB7'>K</span><span style='color:#ffffff'>os</span>
  </div>
  <div style='width:1px;height:22px;background:#30363d'></div>
  <div style='font-size:11px;color:#8b949e;font-style:italic;letter-spacing:0.04em'>
    Support Intelligence Platform
  </div>
</div>
""", unsafe_allow_html=True)

# Customer dropdown + tab buttons in one row
nav_cols = st.columns([1.2] + [0.7] * len(TABS) + [0.1])

with nav_cols[0]:
    selected_customer = st.selectbox(
        "", list(CUSTOMERS.keys()),
        index=list(CUSTOMERS.keys()).index(st.session_state.selected_customer),
        key="customer_select",
        label_visibility="collapsed",
    )
    if selected_customer != st.session_state.selected_customer:
        st.session_state.selected_customer = selected_customer
        st.cache_data.clear()
        st.rerun()

for i, (icon, label) in enumerate(TABS):
    with nav_cols[i + 1]:
        is_active = st.session_state.active_tab == i
        btn_style = "primary" if is_active else "secondary"
        if st.button(f"{icon} {label}", key=f"tab_{i}", type=btn_style,
                     use_container_width=True):
            st.session_state.active_tab = i
            st.rerun()

customer_name = st.session_state.selected_customer
customer      = CUSTOMERS[customer_name]
schema        = customer["schema"]
go_live       = customer.get("go_live", "—")

# ── Customer header strip ─────────────────────────────────────────────────────
st.markdown(f"""
<div style='background:#161b22;border-radius:8px;padding:8px 16px;margin-bottom:12px;
     border:0.5px solid #30363d;display:flex;gap:32px;flex-wrap:wrap;align-items:center'>
  <div style='font-size:15px;font-weight:600;color:#e6edf3'>{customer_name}</div>
  <div style='font-size:11px;color:#8b949e'>Go-live: <b style="color:#c9d1d9">{go_live}</b></div>
  <div style='font-size:11px;color:#8b949e'>CSM: <b style="color:#c9d1d9">{customer.get("csm","—")}</b></div>
  <div style='font-size:11px;color:#8b949e'>License: <b style="color:#c9d1d9">{customer.get("license","—")}</b></div>
  <div style='font-size:11px;color:#8b949e'>Schema: <code style="color:#79c0ff">{schema}</code></div>
</div>
""", unsafe_allow_html=True)

# ── Route to active tab ───────────────────────────────────────────────────────
tab = st.session_state.active_tab

# ════════════════════════════════════════════════════════════════════════════
# TAB 0 — ACCOUNT SNAPSHOT
# ════════════════════════════════════════════════════════════════════════════
if tab == 0:
    from utils.snowflake_conn import run_query
    from utils.charts import COLORS, area_chart, combo_chart, gauge_chart, donut_chart
    from utils.safe_data import safe_float, safe_int, ica_active, ica_counts
    from utils.slack_feed import get_highlights, RISK_STYLE, TYPE_ICON
    from queries.registry import QUERIES

    with st.spinner("Loading..."):
        df_frt    = run_query(QUERIES["frt_monthly"],            schema)
        df_esc    = run_query(QUERIES["escalation_monthly"],     schema)
        df_health = run_query(QUERIES["account_health_monthly"], schema)
        df_sent   = run_query(QUERIES["sentiment_monthly"],      schema)
        df_ica    = run_query(QUERIES["ica_total_lifetime"],     schema)
        df_vol    = run_query(QUERIES["case_volume_ytd"],       schema)
        df_vol_mo = run_query(QUERIES["case_volume_monthly"],   schema)
        df_eng    = run_query(QUERIES["engagement_metrics"],    schema)

    c1,c2,c3,c4,c5 = st.columns(5)
    if not df_frt.empty:
        lf = safe_float(df_frt,"avg_frt_hours",idx=-2)
        ff = safe_float(df_frt,"avg_frt_hours",idx=0)
        c1.metric("First response time", f"{lf} hrs",
                  f"↓ {round((1-lf/ff)*100)}% since go-live" if ff else "—", delta_color="normal")
    if not df_esc.empty:
        le = safe_float(df_esc,"escalation_pct",idx=-2)
        c2.metric("Escalation rate", f"{le}%",
                  "✓ Below 2% benchmark" if le<2 else "⚠ Above benchmark", delta_color="inverse")
    if not df_health.empty:
        lh = safe_float(df_health,"avg_health_score",idx=-1)
        ph = safe_float(df_health,"avg_health_score",idx=-2)
        dh = round(lh-ph,1)
        c3.metric("Account health", f"{lh}/100",
                  f"{'+' if dh>=0 else ''}{dh} vs prior month",
                  delta_color="normal" if dh>=0 else "inverse")
    if not df_sent.empty:
        c4.metric("Avg sentiment", f"{safe_float(df_sent,'avg_sentiment',idx=-1)}/100","Stable")
    auto_cases, manual_cases = ica_counts(df_ica)
    if ica_active(df_ica):
        c5.metric("ICA auto-assignments", f"{auto_cases:,}", "Lifetime total")
    else:
        c5.metric("ICA","Not deployed","Manual assignment active", delta_color="off")

    # ── Case volume row ───────────────────────────────────────────────────────
    if not df_vol.empty:
        import datetime
        curr_yr = datetime.datetime.now().year
        v1,v2,v3,v4 = st.columns(4)
        ytd   = int(df_vol["ytd_cases"].iloc[0])
        avg_m = int(df_vol["avg_monthly"].iloc[0])
        cur_m = int(df_vol["current_month_cases"].iloc[0])
        prv_m = int(df_vol["prev_month_cases"].iloc[0])
        m_delta = cur_m - prv_m
        v1.metric(f"Cases created YTD ({curr_yr})",
                  f"{ytd:,}",
                  "Excl. bots & deleted",
                  delta_color="off")
        v2.metric("Avg monthly inflow",
                  f"{avg_m:,}",
                  f"Based on {int(df_vol['months_with_data'].iloc[0])} months YTD",
                  delta_color="off")
        v3.metric("This month (so far)",
                  f"{cur_m:,}",
                  f"{'+' if m_delta>=0 else ''}{m_delta:,} vs last month",
                  delta_color="inverse" if m_delta>avg_m*0.1 else "off")
        v4.metric("Last full month",
                  f"{prv_m:,}",
                  "Complete month",
                  delta_color="off")
        st.caption(
            f"Case volume: Jan {curr_yr} – present · "
            "Excludes bot cases (sl_is_bot=false) and deleted cases (s_deleted_at IS NULL)"
        )

    # ── Engagement metrics ────────────────────────────────────────────────────
    if not df_eng.empty:
        import datetime
        from datetime import date
        curr_yr    = datetime.datetime.now().year
        today      = date.today()
        day_of_year = today.timetuple().tm_yday
        ann_factor  = 365 / day_of_year

        st.markdown("#### 📊 Annual engagement summary")
        st.caption(
            f"Jan {curr_yr} to present ({day_of_year} days elapsed) · "
            f"Projected annual = YTD × {ann_factor:.1f} · All SupportLogic modules"
        )

        def _fmt(v):
            try:
                n = int(float(v))
                if n >= 1000000: return f"{n/1000000:.1f}M"
                if n >= 1000:    return f"{n/1000:.0f}K"
                return f"{n:,}"
            except: return "—"

        def _proj(v):
            try:
                n = int(float(v))
                proj = int(n * ann_factor)
                return _fmt(n), f"Proj. annual: {_fmt(proj)}"
            except: return "—", ""

        row = df_eng.iloc[0]

        # Row 1 — Core processing
        e1,e2,e3,e4 = st.columns(4)
        v,s = _proj(row.get("annual_interactions",0))
        e1.metric("Interactions processed", v, s, delta_color="off")
        v,s = _proj(row.get("annual_alerts",0))
        e2.metric("Alerts fired (YTD)", v, s, delta_color="off")
        v,s = _proj(row.get("annual_signals",0))
        e3.metric("Signals extracted", v, s, delta_color="off")
        v,s = _proj(row.get("annual_summaries",0))
        e4.metric("AI summaries generated", v, s, delta_color="off")

        # Row 2 — Feature metrics
        e5,e6,e7,e8 = st.columns(4)
        ica_val = int(float(row.get("projected_annual_ica",0) or 0))
        if ica_val > 0:
            e5.metric("ICA assignments (proj.)", _fmt(ica_val), "Last full month × 12", delta_color="off")
        else:
            e5.metric("ICA assignments", "Not deployed", "ICA not active", delta_color="off")

        v,s = _proj(row.get("annual_auto_qa",0))
        e6.metric("Auto QA performed",
                  v if v not in ("0","—") else "Not active",
                  s if v not in ("0","—") else "Elevate not deployed",
                  delta_color="off")

        # Resolve API from XFIND DB — YTD with projection
        xfind_schemas = customer.get("xfind_schemas", [])
        if xfind_schemas:
            try:
                xfind_ytd = 0
                for xs in xfind_schemas:
                    _xf = run_query(
                        f"""SELECT COUNT(*) AS cnt
                            FROM XFIND.{xs}.CORE_QUERYACTIVITY
                            WHERE CREATED_AT >= DATE_TRUNC('YEAR', CURRENT_DATE())""",
                        schema
                    )
                    xfind_ytd += int(_xf["cnt"].iloc[0])
                xfind_proj = int(xfind_ytd * ann_factor)
                e7.metric("Resolve API calls (YTD)", f"{_fmt(xfind_ytd)}",
                          f"Proj. annual: {_fmt(xfind_proj)}", delta_color="off")
            except Exception as _xe:
                e7.metric("Resolve API calls", "Error", str(_xe)[:40], delta_color="off")
        else:
            e7.metric("Resolve API calls", "Not deployed",
                      "xFind not configured", delta_color="off")

        # Iframe MAU from Pendo
        try:
            from utils.pendo_conn import get_page_views, get_combined_page_views
            _pids = customer.get("pendo_ids", [customer.get("pendo_id", customer_name.lower())])
            if len(_pids) > 1:
                _dp = get_combined_page_views(_pids, 30)
            else:
                _dp = get_page_views(_pids[0], 30)
            # Filter to iframe pages only
            if not _dp.empty and "pageId" in _dp.columns:
                _page_map = get_page_map() if "get_page_map" in dir() else {}
                _iframe_df = _dp[_dp["pageId"].map(
                    lambda x: "[IFrame]" in str(_page_map.get(x, x))
                )] if _page_map else _dp
                _imau = _iframe_df["visitorId"].nunique() if not _iframe_df.empty else 0
            else:
                _imau = 0
            e8.metric("Iframe MAU (last 30d)", f"{_imau:,}",
                      "Agents via CRM iframe only · Pendo", delta_color="off")
        except Exception:
            e8.metric("Iframe MAU", "—", "Pendo not configured", delta_color="off")

        # Row 3 — Dashboard MAU + Agent count
        e9,e10,_,__ = st.columns(4)
        try:
            from utils.pendo_conn import get_page_views, get_combined_visitor_count
            _pids2 = customer.get("pendo_ids", [customer.get("pendo_id", customer_name.lower())])
            if len(_pids2) > 1:
                _dmau = get_combined_visitor_count(_pids2, 90)
            else:
                _dp90 = get_page_views(_pids2[0], 90)
                _dmau = _dp90["visitorId"].nunique() if not _dp90.empty and "visitorId" in _dp90.columns else 0
            e9.metric("Active UI users (90d)", f"{_dmau:,}",
                      "Unique users who opened SL UI · last 90d", delta_color="off")
        except Exception:
            e9.metric("Dashboard visitors (90d)", "—", "Pendo not configured", delta_color="off")

        try:
            _ag = run_query("""
                SELECT COUNT(DISTINCT sl_assignee_id) AS agent_count
                FROM PIPE_DATABASE.<SCHEMA>.case_summary
                WHERE sl_created_at >= DATEADD(DAY,-60,CURRENT_DATE())
                  AND sl_is_bot = FALSE
                  AND is_deleted = FALSE
                  AND sl_assignee_id IS NOT NULL
            """, schema)
            _agc = int(_ag["agent_count"].iloc[0])
            e10.metric("Unique assignees (60d)", f"{_agc:,}",
                       "Unique users cases were assigned to · last 60d", delta_color="off")
        except Exception:
            e10.metric("Est. active agents", "—", "Query failed", delta_color="off")
    st.divider()
    # ── Monthly volume chart ─────────────────────────────────────────────────
    if not df_vol_mo.empty:
        from utils.charts import stacked_bar, line_chart
        col_vol, col_frt2 = st.columns(2)
        with col_vol:
            st.markdown("**Monthly case inflow — YTD (excl. bots & deleted)**")
            st.plotly_chart(stacked_bar(df_vol_mo, x="month",
                y_cols=[
                    {"col":"p1_cases","name":"P1 Critical","color":COLORS["red"]},
                    {"col":"p2_cases","name":"P2 High",    "color":COLORS["amber"]},
                    {"col":"total_cases","name":"Total",   "color":COLORS["blue_light"]},
                ], height=220), use_container_width=True)
        with col_frt2:
            if not df_frt.empty:
                st.markdown("**First response time — monthly (hrs)**")
                st.plotly_chart(area_chart(df_frt, x="month",
                    y_cols=[{"col":"avg_frt_hours","name":"FRT (hrs)",
                             "color":COLORS["blue"],"fill":"rgba(24,95,165,0.1)"}],
                    height=220), use_container_width=True)
        st.divider()

    col_g1,col_g2,col_g3,col_area = st.columns([1,1,1,3])
    if not df_health.empty:
        with col_g1:
            st.markdown("**Health score**")
            st.plotly_chart(gauge_chart(safe_float(df_health,"avg_health_score",idx=-1),label="/ 100"),use_container_width=True)
    if not df_sent.empty:
        with col_g2:
            st.markdown("**Sentiment**")
            st.plotly_chart(gauge_chart(safe_float(df_sent,"avg_sentiment",idx=-1),label="/ 100"),use_container_width=True)
    if not df_esc.empty:
        le=safe_float(df_esc,"escalation_pct",idx=-2)
        with col_g3:
            st.markdown("**Esc. rate**")
            st.plotly_chart(gauge_chart(le,min_val=0,max_val=5,
                thresholds=[{"range":[0,1],"color":COLORS["teal_light"]},
                            {"range":[1,2],"color":COLORS["amber_light"]},
                            {"range":[2,5],"color":COLORS["red_light"]}],label="%"),use_container_width=True)
    if not df_frt.empty:
        with col_area:
            st.markdown("**First response time — monthly (hrs)**")
            st.plotly_chart(area_chart(df_frt,x="month",
                y_cols=[{"col":"avg_frt_hours","name":"FRT (hrs)","color":COLORS["blue"],"fill":"rgba(24,95,165,0.1)"}],
                height=200),use_container_width=True)

    st.divider()
    col_d,col_c = st.columns([1,2])
    if ica_active(df_ica):
        total = auto_cases+manual_cases
        with col_d:
            st.markdown("**ICA auto vs manual (lifetime)**")
            st.plotly_chart(donut_chart(
                labels=["Auto (ICA)","Manual"],values=[auto_cases,manual_cases],
                colors=[COLORS["teal"],COLORS["gray"]],height=240,
                center_text=f"{round(auto_cases/total*100) if total else 0}% auto"),use_container_width=True)
    else:
        with col_d:
            st.markdown("**ICA — Routing Agent**")
            st.info("ICA not deployed for this customer.",icon="ℹ️")
    if not df_esc.empty:
        with col_c:
            st.markdown("**Escalation rate vs case volume**")
            st.plotly_chart(combo_chart(df_esc,x="month",
                bar_col={"col":"total_cases","name":"Cases","color":COLORS["gray_light"]},
                line_col={"col":"escalation_pct","name":"Esc %","color":COLORS["red"]},
                reference_lines=[{"y":2.0,"label":"2% benchmark","color":COLORS["red"]}],
                height=240),use_container_width=True)

    col_h,col_s = st.columns(2)
    if not df_health.empty:
        with col_h:
            st.markdown("**Account health score trend**")
            st.plotly_chart(area_chart(df_health,x="month",
                y_cols=[{"col":"avg_health_score","name":"Health score","color":COLORS["purple"],"fill":"rgba(83,74,183,0.1)"}],
                reference_lines=[{"y":80,"label":"Score 80","color":COLORS["gray"]}],height=220),use_container_width=True)
    if not df_sent.empty:
        with col_s:
            st.markdown("**Sentiment vs need-attention**")
            st.plotly_chart(area_chart(df_sent,x="month",
                y_cols=[
                    {"col":"avg_sentiment","name":"Sentiment","color":COLORS["teal"],"fill":"rgba(15,110,86,0.1)"},
                    {"col":"avg_need_attention","name":"Need attention","color":COLORS["amber"],"fill":"rgba(186,117,23,0.08)"},
                ],height=220),use_container_width=True)

    st.divider()
    st.subheader("📣 Recent account activity")
    highlights = get_highlights(customer_name)
    if not highlights:
        st.info(f"No Slack highlights loaded for {customer_name}.",icon="ℹ️")
    else:
        for h in highlights:
            icon,bg,bc = RISK_STYLE.get(h["risk"],RISK_STYLE["low"])
            ti = TYPE_ICON.get(h["type"],"📌")
            acts = "".join(f"<li style='margin:2px 0'>{a}</li>" for a in h.get("actions",[]))
            acts_html = f"<ul style='margin:6px 0 0;padding-left:16px;font-size:11px;color:#8b949e'>{acts}</ul>" if acts else ""
            st.markdown(f"""<div style='background:{bg};border:0.5px solid {bc};border-radius:8px;padding:12px 16px;margin-bottom:8px'>
              <div style='display:flex;justify-content:space-between'>
                <div style='font-size:13px;font-weight:500;color:#e6edf3'>{ti} {h["title"]}</div>
                <div style='font-size:11px;color:#8b949e'>{icon} {h["date"]}</div>
              </div>
              <div style='font-size:12px;color:#c9d1d9;margin-top:4px;line-height:1.5'>{h["summary"]}</div>
              {acts_html}
            </div>""",unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════════════════
# TAB 1 — ESCALATION
# ════════════════════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════════════════════
# TAB 1 — ACCOUNT INTELLIGENCE BRIEF
# ════════════════════════════════════════════════════════════════════════════
elif tab == 1:
    from utils.brief_generator import generate_brief, is_available
    from utils.snowflake_conn import run_query
    from queries.registry import QUERIES
    from utils.pendo_conn import (get_page_views, get_feature_events,
        get_page_map, get_feature_map, get_visitor_count,
        dau, top_modules, top_features)

    SECTION_ICONS = {
        "account_status":   ("🔵", "Account Status"),
        "value_delivered":  ("📈", "Value Delivered"),
        "platform_adoption":("👥", "Platform Adoption"),
        "working_well":     ("✅", "Working Well"),
        "watch_items":      ("⚠️",  "Watch Items"),
        "renewal_outlook":  ("🔄", "Renewal Outlook"),
    }

    SECTION_BG = {
        "account_status":   "#161b22",
        "value_delivered":  "#0f2a1a",
        "platform_adoption":"#0a1931",
        "working_well":     "#0f2a1a",
        "watch_items":      "#2a1010",
        "renewal_outlook":  "#1a1428",
    }

    SECTION_BORDER = {
        "account_status":   "#30363d",
        "value_delivered":  "#1D9E75",
        "platform_adoption":"#185FA5",
        "working_well":     "#1D9E75",
        "watch_items":      "#E24B4A",
        "renewal_outlook":  "#534AB7",
    }

    # Check if LLM available
    if not is_available():
        st.markdown("""
<div style='background:#1e1e1e;border:0.5px solid #444;border-radius:10px;
     padding:24px 28px;text-align:center;margin-top:40px'>
  <div style='font-size:32px;margin-bottom:12px'>🧠</div>
  <div style='font-size:16px;font-weight:600;color:#e6edf3;margin-bottom:8px'>
    Account Intelligence Brief — Unavailable
  </div>
  <div style='font-size:13px;color:#8b949e;line-height:1.6'>
    This feature requires an Anthropic API key.<br>
    Add <code>ANTHROPIC_API_KEY=your_key</code> to your <code>.env</code> file<br>
    or Streamlit Cloud secrets to enable AI-generated account briefs.
  </div>
</div>""", unsafe_allow_html=True)
    else:
        st.markdown("""
<div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:4px'>
  <div>
    <div style='font-size:20px;font-weight:700;color:#e6edf3'>🧠 Account Intelligence Brief</div>
    <div style='font-size:12px;color:#8b949e;margin-top:2px'>
      AI-generated summary from live Snowflake + Pendo data · Manual trigger · Factual only
    </div>
  </div>
</div>""", unsafe_allow_html=True)

        brief_key = f"brief_{customer_name}"

        col_btn, col_info = st.columns([1, 3])
        with col_btn:
            generate = st.button("🧠 Generate brief", type="primary",
                                 use_container_width=True)
        with col_info:
            if brief_key in st.session_state:
                b = st.session_state[brief_key]
                st.caption(f"Generated: {b.get('generated_at','—')} · "
                           f"Period: {b['facts'].get('period_label','—')} · "
                           f"Switch customer to auto-clear")
            else:
                st.caption("Pulls from live Snowflake + Pendo data · "
                           "~15 seconds · Uses Claude API once per generation")

        if generate:
            with st.spinner("Loading data and generating brief..."):
                brief_data = {
                    "frt_monthly":          run_query(QUERIES["frt_monthly"],          schema),
                    "escalation_monthly":   run_query(QUERIES["escalation_monthly"],   schema),
                    "lte_accuracy_monthly": run_query(QUERIES["lte_accuracy_monthly"], schema),
                    "reassignment_monthly": run_query(QUERIES["reassignment_monthly"], schema),
                    "account_health_monthly":run_query(QUERIES["account_health_monthly"],schema),
                    "sentiment_monthly":    run_query(QUERIES["sentiment_monthly"],    schema),
                    "alerts_monthly":       run_query(QUERIES["alerts_monthly"],       schema),
                    "ica_total_lifetime":   run_query(QUERIES["ica_total_lifetime"],   schema),
                    "ai_summaries_total":   run_query(QUERIES["ai_summaries_total"],   schema),
                    "platform_actions_summary":run_query(QUERIES["platform_actions_summary"],schema),
                }
                pendo_brief = None
                try:
                    _pid = customer.get("pendo_id", customer_name.lower())
                    _pm  = get_page_map(); _fm = get_feature_map()
                    _dp  = get_page_views(_pid, 30)
                    _df  = get_feature_events(_pid, 30)
                    _dau = dau(_dp)
                    import pandas as _pd
                    _pu  = _dp.groupby("visitorId")["numMinutes"].sum().reset_index() if not _dp.empty else _pd.DataFrame()
                    _tot = len(_pu)
                    _a   = len(_pu[_pu["numMinutes"]>=180]) if not _pu.empty else 0
                    _mo  = len(_pu[(_pu["numMinutes"]>=60)&(_pu["numMinutes"]<180)]) if not _pu.empty else 0
                    _lo  = len(_pu[_pu["numMinutes"]<60])  if not _pu.empty else 0
                    _p   = lambda n: round(n/_tot*100) if _tot else 0
                    pendo_brief = {
                        "visitors": get_visitor_count(_pid),
                        "avg_dau":  round(float(_dau["dau"].mean())) if not _dau.empty else 0,
                        "days": 30,
                        "tiers": {"active":_a,"moderate":_mo,"low":_lo,
                                  "active_pct":_p(_a),"moderate_pct":_p(_mo),
                                  "low_pct":_p(_lo),"total":_tot},
                        "top_modules": top_modules(_dp, _pm, n=1),
                    }
                except Exception:
                    pass

                brief = generate_brief(customer_name, customer, brief_data, pendo_brief)
                st.session_state[brief_key] = brief

        # Render brief if available
        if brief_key in st.session_state:
            b = st.session_state[brief_key]
            st.divider()

            for key, (icon, label) in SECTION_ICONS.items():
                content = b.get(key, "—")
                if content and content != "—":
                    bg     = SECTION_BG.get(key, "#161b22")
                    border = SECTION_BORDER.get(key, "#30363d")
                    st.markdown(f"""
<div style='background:{bg};border:0.5px solid {border};border-radius:10px;
     padding:16px 20px;margin-bottom:10px'>
  <div style='font-size:11px;font-weight:700;color:{border};letter-spacing:.06em;
       text-transform:uppercase;margin-bottom:8px'>{icon} {label}</div>
  <div style='font-size:13px;color:#c9d1d9;line-height:1.7'>{content}</div>
</div>""", unsafe_allow_html=True)

            # Metadata footer
            f_ = b.get("facts", {})
            st.markdown(f"""
<div style='background:#0d1117;border:0.5px solid #21262d;border-radius:8px;
     padding:10px 16px;margin-top:8px;font-size:10px;color:#484f58;line-height:1.8'>
  <b style='color:#30363d'>Data sources:</b>
  Snowflake PIPE_DATABASE.{schema} · Pendo (last 30d) · Generated {b.get('generated_at','—')} ·
  Period: {f_.get('period_label','—')} · Cases: {f_.get('total_cases',0):,} ·
  Benchmark escalation rate: {f_.get('benchmark',2.0)}%
</div>""", unsafe_allow_html=True)
        elif not generate:
            st.markdown("""
<div style='text-align:center;padding:60px 20px;color:#8b949e'>
  <div style='font-size:40px;margin-bottom:12px'>🧠</div>
  <div style='font-size:14px'>Click <b>Generate brief</b> to create an AI account summary</div>
  <div style='font-size:12px;margin-top:6px;color:#484f58'>
    Pulls live data from Snowflake and Pendo · Takes ~15 seconds · Uses ~1,500 tokens
  </div>
</div>""", unsafe_allow_html=True)


elif tab == 2:
    from utils.snowflake_conn import run_query
    from utils.charts import COLORS, line_chart, combo_chart, stacked_bar
    from queries.registry import QUERIES
    from utils.safe_data import safe_float, safe_int

    benchmark = customer.get("benchmark_escalation_pct",2.0)
    with st.spinner("Loading..."):
        df_esc = run_query(QUERIES["escalation_monthly"],schema)
        df_lte = run_query(QUERIES["lte_accuracy_monthly"],schema)
        df_rev = run_query(QUERIES["escalation_reviews_monthly"],schema)

    c1,c2,c3,c4 = st.columns(4)
    if not df_lte.empty:
        pred=int(df_lte["cases_predicted"].sum()); act=int(df_lte["actually_escalated"].sum()); prev=pred-act
        c1.metric("Cases flagged by LTE",f"{pred:,}")
        c2.metric("Escalations prevented",f"{prev:,}",f"{round(prev/pred*100) if pred else 0}% intercepted",delta_color="normal")
        c3.metric("Avg LTE hit rate",f"{round(float(df_lte['hit_rate_pct'].mean()),1)}%","Low = agents intervening ✓",delta_color="off")
    if not df_esc.empty:
        le=safe_float(df_esc,"escalation_pct",idx=-2)
        c4.metric("Latest esc. rate",f"{le}%",f"{'✓ Below' if le<benchmark else '⚠ Above'} {benchmark}% benchmark",
                  delta_color="normal" if le<benchmark else "inverse")

    st.info("**LTE hit rate:** Low (3–8%) means SupportLogic flags risk early enough for agents to intervene before formal escalation.",icon="ℹ️")
    st.divider()

    col_l,col_r = st.columns(2)
    if not df_lte.empty:
        with col_l:
            st.markdown("**LTE predictions vs actual escalations**")
            st.plotly_chart(combo_chart(df_lte,x="month",
                bar_col={"col":"cases_predicted","name":"Predicted (LTE)","color":COLORS["amber_light"]},
                line_col={"col":"actually_escalated","name":"Actually escalated","color":COLORS["red"]},
                height=260),use_container_width=True)
    if not df_esc.empty:
        with col_r:
            st.markdown("**Escalation rate vs benchmark**")
            st.plotly_chart(line_chart(df_esc,x="month",
                y_cols=[{"col":"escalation_pct","name":"Esc %","color":COLORS["red"]}],
                reference_lines=[{"y":benchmark,"label":f"{benchmark}% benchmark","color":COLORS["red"]}],
                height=260),use_container_width=True)

    col_l2,col_r2 = st.columns(2)
    if not df_lte.empty:
        with col_l2:
            st.markdown("**LTE hit rate % — monthly**")
            st.plotly_chart(line_chart(df_lte,x="month",
                y_cols=[{"col":"hit_rate_pct","name":"Hit rate %","color":COLORS["purple"]}],
                height=240),use_container_width=True)
    if not df_rev.empty:
        with col_r2:
            pivot=df_rev.pivot_table(index="month",columns="status",values="cnt",aggfunc="sum",fill_value=0).reset_index()
            sc={"COMPLETED":COLORS["teal"],"PENDING":COLORS["amber"],"SNOOZED":COLORS["blue"],"REJECTED":COLORS["red"]}
            y_cols=[{"col":s,"name":s.capitalize(),"color":sc.get(s,COLORS["gray"])} for s in [c for c in pivot.columns if c!="month"]]
            st.markdown("**Escalation review swarming — monthly**")
            st.plotly_chart(stacked_bar(pivot,x="month",y_cols=y_cols,height=240),use_container_width=True)

# ════════════════════════════════════════════════════════════════════════════
# TAB 2 — SENTIMENT
# ════════════════════════════════════════════════════════════════════════════
elif tab == 3:
    from utils.snowflake_conn import run_query
    from utils.charts import COLORS, area_chart, stacked_bar
    from queries.registry import QUERIES
    from utils.safe_data import safe_float, safe_int

    with st.spinner("Loading..."):
        df_sent   = run_query(QUERIES["sentiment_monthly"],schema)
        df_alerts = run_query(QUERIES["alerts_monthly"],schema)

    c1,c2,c3,c4 = st.columns(4)
    if not df_sent.empty:
        c1.metric("Avg sentiment (latest)",f"{safe_float(df_sent,'avg_sentiment',idx=-1)}/100")
        c2.metric("Need-attention score",f"{safe_float(df_sent,'avg_need_attention',idx=-1)}/100","Lower = healthier",delta_color="off")
        c3.metric("Cases scored (latest)",f"{int(df_sent['cases_scored'].iloc[-1]):,}","100% coverage")
    if not df_alerts.empty:
        c4.metric("Total alerts (period)",f"{int(df_alerts['total_alerts'].sum()):,}")

    st.divider()
    col_l,col_r = st.columns(2)
    if not df_sent.empty:
        with col_l:
            st.markdown("**Sentiment & need-attention — monthly**")
            st.plotly_chart(area_chart(df_sent,x="month",
                y_cols=[
                    {"col":"avg_sentiment","name":"Sentiment","color":COLORS["teal"],"fill":"rgba(15,110,86,0.12)"},
                    {"col":"avg_need_attention","name":"Need-attention","color":COLORS["amber"],"fill":"rgba(186,117,23,0.08)"},
                ],height=280),use_container_width=True)
    if not df_alerts.empty:
        with col_r:
            st.markdown("**Alerts consumption — monthly**")
            st.plotly_chart(stacked_bar(df_alerts,x="month",
                y_cols=[
                    {"col":"alert_cases","name":"Alert cases","color":COLORS["teal"]},
                    {"col":"total_alerts","name":"Total alerts","color":COLORS["teal_light"]},
                ],height=280),use_container_width=True)

# ════════════════════════════════════════════════════════════════════════════
# TAB 3 — ROUTING
# ════════════════════════════════════════════════════════════════════════════
elif tab == 4:
    from utils.snowflake_conn import run_query
    from utils.charts import COLORS, line_chart, bar_chart, stacked_bar
    from utils.safe_data import safe_float, safe_int, ica_active, ica_counts
    from queries.registry import QUERIES

    with st.spinner("Loading..."):
        df_ica    = run_query(QUERIES["ica_total_lifetime"],schema)
        df_ica_mo = run_query(QUERIES["ica_events_monthly"],schema)
        df_re     = run_query(QUERIES["reassignment_monthly"],schema)
        df_fp     = run_query(QUERIES["frt_by_priority"],schema)
        df_cpa    = run_query(QUERIES["cases_per_agent_weekly"],schema)
        df_frt    = run_query(QUERIES["frt_monthly"],schema)

    auto_cases,manual_cases = ica_counts(df_ica)
    ica_on = ica_active(df_ica)

    c1,c2,c3,c4 = st.columns(4)
    if ica_on:
        total=auto_cases+manual_cases
        c1.metric("ICA auto-assignments",f"{auto_cases:,}")
        c2.metric("Auto rate",f"{round(auto_cases/total*100) if total else 0}%",f"{manual_cases:,} manual",delta_color="off")
    else:
        c1.metric("ICA","Not deployed","No auto-routing",delta_color="off")
        c2.metric("Manual assignments",f"{manual_cases:,}" if manual_cases else "—")
    if not df_re.empty:
        fr=safe_float(df_re,"reassignment_pct",idx=0); lr=safe_float(df_re,"reassignment_pct",idx=-2)
        c3.metric("Reassignment rate",f"{lr}%",f"↓ {round(fr-lr,1)}pp from {fr}%",delta_color="normal")
    if not df_frt.empty:
        c4.metric("Overall FRT (latest)",f"{safe_float(df_frt,'avg_frt_hours',idx=-2)} hrs")

    st.divider()
    if ica_on and not df_ica_mo.empty:
        st.markdown("**ICA auto vs manual assignments — monthly**")
        st.plotly_chart(stacked_bar(df_ica_mo,x="month",
            y_cols=[{"col":"auto_cases","name":"Auto (ICA)","color":COLORS["teal"]},
                    {"col":"manual_cases","name":"Manual","color":COLORS["gray"]}],height=220),use_container_width=True)
        st.divider()
    else:
        st.info("**ICA Routing Agent** — not deployed for this customer.",icon="ℹ️")

    col_l,col_r = st.columns(2)
    if not df_re.empty:
        clrs=[COLORS["red"] if v>5 else COLORS["amber"] if v>1 else COLORS["teal"] for v in df_re["reassignment_pct"]]
        with col_l:
            st.markdown("**Reassignment rate — monthly**")
            st.plotly_chart(bar_chart(df_re,x="month",
                y_cols=[{"col":"reassignment_pct","name":"Reassignment %","color":clrs,"colors":clrs}],
                height=260),use_container_width=True)
    if not df_fp.empty:
        pivot=df_fp.pivot_table(index="month",columns="priority_tier",values="avg_frt_hours",aggfunc="mean").reset_index()
        cols=[c for c in ["P1 Critical","P2 High","P3 Medium","P4/Other"] if c in pivot.columns]
        cm={"P1 Critical":COLORS["red"],"P2 High":COLORS["amber"],"P3 Medium":COLORS["blue"],"P4/Other":COLORS["gray"]}
        with col_r:
            st.markdown("**FRT by priority tier (hrs)**")
            st.plotly_chart(line_chart(pivot,x="month",
                y_cols=[{"col":c,"name":c,"color":cm.get(c,COLORS["gray"])} for c in cols],
                height=260),use_container_width=True)

    col_l2,col_r2 = st.columns(2)
    if not df_cpa.empty:
        df_plot=df_cpa.iloc[:-1].copy()
        df_plot["week_start"]=pd.to_datetime(df_plot["week_start"])
        df_plot=df_plot[df_plot["week_start"].dt.year>=2025]
        df_plot["week_label"]=df_plot["week_start"].dt.strftime("%-m/%-d")
        with col_l2:
            st.markdown("**Cases per agent — weekly**")
            st.plotly_chart(bar_chart(df_plot,x="week_label",
                y_cols=[{"col":"cases_per_agent","name":"Cases/agent","color":COLORS["purple"]}],
                height=240),use_container_width=True)
    if not df_frt.empty:
        with col_r2:
            st.markdown("**Overall FRT — monthly**")
            st.plotly_chart(line_chart(df_frt,x="month",
                y_cols=[{"col":"avg_frt_hours","name":"FRT (hrs)","color":COLORS["blue"]}],
                height=240),use_container_width=True)

# ════════════════════════════════════════════════════════════════════════════
# TAB 4 — FEATURE USAGE
# ════════════════════════════════════════════════════════════════════════════
elif tab == 5:
    from utils.snowflake_conn import run_query
    from utils.charts import COLORS, stacked_bar, donut_chart, bar_chart
    from utils.feature_flags import detect_features
    from utils.safe_data import safe_int, ica_active, ica_counts
    from queries.registry import QUERIES

    with st.spinner("Loading..."):
        data = {
            "ica_total_lifetime":         run_query(QUERIES["ica_total_lifetime"],schema),
            "ica_events_monthly":         run_query(QUERIES["ica_events_monthly"],schema),
            "lte_accuracy_monthly":       run_query(QUERIES["lte_accuracy_monthly"],schema),
            "sentiment_monthly":          run_query(QUERIES["sentiment_monthly"],schema),
            "account_health_monthly":     run_query(QUERIES["account_health_monthly"],schema),
            "alerts_monthly":             run_query(QUERIES["alerts_monthly"],schema),
            "ai_summaries_total":         run_query(QUERIES["ai_summaries_total"],schema),
            "ai_summaries_monthly":       run_query(QUERIES["ai_summaries_monthly"],schema),
            "escalation_reviews_monthly": run_query(QUERIES["escalation_reviews_monthly"],schema),
            "platform_actions_summary":   run_query(QUERIES["platform_actions_summary"],schema),
            "platform_actions_monthly":   run_query(QUERIES["platform_actions_monthly"],schema),
        }

    features = detect_features(data)
    st.subheader("Feature activation status")
    st.caption("Derived from live data — not hardcoded.")

    icon_m={"active":"✅","not_deployed":"🔴","not_licensed":"⬜"}
    style_m={"active":"background:#1a3a2a;border:0.5px solid #1D9E75",
              "not_deployed":"background:#3a1a1a;border:0.5px solid #E24B4A",
              "not_licensed":"background:#1e1e1e;border:0.5px solid #444"}
    label_m={"active":"Active","not_deployed":"Not deployed / no activity","not_licensed":"Not licensed"}

    cols=st.columns(3)
    for i,f in enumerate(features):
        with cols[i%3]:
            st.markdown(f"""<div style='padding:10px 14px;border-radius:8px;{style_m[f["status"]]};margin-bottom:8px'>
              <div style='font-size:13px;font-weight:500;color:#e6edf3'>{icon_m[f["status"]]} {f["name"]}
                <span style='font-size:10px;font-weight:400;color:#8b949e;margin-left:6px'>{label_m[f["status"]]}</span>
              </div>
              <div style='font-size:11px;color:#8b949e;margin-top:3px'>{f["evidence"]}</div>
            </div>""",unsafe_allow_html=True)

    st.divider()
    df_act_sum = data["platform_actions_summary"]
    df_act_mo  = data["platform_actions_monthly"]
    if not df_act_sum.empty and int(df_act_sum["total_actions"].sum())>0:
        col_l,col_r = st.columns(2)
        palette=[COLORS["teal"],COLORS["blue"],COLORS["purple"],COLORS["amber"],
                 COLORS["red"],COLORS["green"],COLORS["coral"],COLORS["gray"]]
        with col_l:
            clrs=[palette[i%len(palette)] for i in range(len(df_act_sum))]
            st.markdown("**Agent actions on SL signals (lifetime)**")
            st.plotly_chart(bar_chart(df_act_sum,x="action_type",
                y_cols=[{"col":"total_actions","name":"Actions","color":clrs,"colors":clrs}],
                height=260,horizontal=True),use_container_width=True)
        if not df_act_mo.empty:
            with col_r:
                pivot=df_act_mo.pivot_table(index="month",columns="action_type",values="cnt",aggfunc="sum",fill_value=0).reset_index()
                act_cols=[c for c in pivot.columns if c!="month"]
                y_cols=[{"col":c,"name":c.replace("_"," ").title(),"color":palette[i%len(palette)]} for i,c in enumerate(act_cols)]
                st.markdown("**Actions by type — monthly**")
                st.plotly_chart(stacked_bar(pivot,x="month",y_cols=y_cols,height=260),use_container_width=True)

    df_st=data["ai_summaries_total"]; df_sm=data["ai_summaries_monthly"]
    if not df_st.empty and int(df_st["total"].sum())>0:
        st.divider()
        col_l2,col_r2=st.columns(2)
        with col_l2:
            st.markdown("**AI summaries — by type (lifetime)**")
            st.plotly_chart(donut_chart(
                labels=df_st["summary_type"].tolist(),values=df_st["total"].tolist(),
                colors=[COLORS["blue"],COLORS["teal"],COLORS["purple"],COLORS["amber"]],
                height=260,center_text=f"{int(df_st['total'].sum()):,} total"),use_container_width=True)
        if not df_sm.empty:
            with col_r2:
                p2=df_sm.groupby(["month","summary_type"])["summaries_generated"].sum().reset_index()
                p2=p2.pivot_table(index="month",columns="summary_type",values="summaries_generated",aggfunc="sum",fill_value=0).reset_index()
                sc=[c for c in p2.columns if c!="month"]
                cmap=[COLORS["blue"],COLORS["teal"],COLORS["purple"],COLORS["amber"]]
                st.markdown("**AI summaries — monthly trend**")
                st.plotly_chart(stacked_bar(p2,x="month",
                    y_cols=[{"col":c,"name":c.replace("_"," ").title(),"color":cmap[i%len(cmap)]} for i,c in enumerate(sc)],
                    height=260),use_container_width=True)

# ════════════════════════════════════════════════════════════════════════════
# TAB 5 — ROI SUMMARY
# ════════════════════════════════════════════════════════════════════════════
elif tab == 6:
    from utils.snowflake_conn import run_query
    from utils.charts import COLORS, area_chart, donut_chart, gauge_chart
    from utils.safe_data import safe_float, safe_int, ica_active, ica_counts
    from queries.registry import QUERIES
    import os, tempfile

    benchmark=customer.get("benchmark_escalation_pct",2.0)
    goals=customer.get("goals",[])

    st.markdown(f"""<div style='background:#161b22;border-radius:10px;padding:14px 18px;margin-bottom:16px;border:0.5px solid #30363d'>
      <div style='display:flex;gap:32px;flex-wrap:wrap'>
        <div><div style='font-size:11px;color:#8b949e'>Customer</div><div style='font-size:14px;font-weight:500;color:#e6edf3'>{customer_name}</div></div>
        <div><div style='font-size:11px;color:#8b949e'>Go-live</div><div style='font-size:14px;font-weight:500;color:#e6edf3'>{go_live}</div></div>
        <div><div style='font-size:11px;color:#8b949e'>CSM</div><div style='font-size:14px;color:#e6edf3'>{customer.get('csm','—')}</div></div>
        <div><div style='font-size:11px;color:#8b949e'>License</div><div style='font-size:14px;color:#e6edf3'>{customer.get('license','—')}</div></div>
      </div>
      {"<div style='margin-top:8px;font-size:12px;color:#8b949e'><b>Goals:</b> "+" · ".join(goals)+"</div>" if goals else ""}
    </div>""",unsafe_allow_html=True)

    with st.spinner("Compiling ROI data..."):
        df_frt     = run_query(QUERIES["frt_monthly"],schema)
        df_esc     = run_query(QUERIES["escalation_monthly"],schema)
        df_lte     = run_query(QUERIES["lte_accuracy_monthly"],schema)
        df_re      = run_query(QUERIES["reassignment_monthly"],schema)
        df_health  = run_query(QUERIES["account_health_monthly"],schema)
        df_sent    = run_query(QUERIES["sentiment_monthly"],schema)
        df_alerts  = run_query(QUERIES["alerts_monthly"],schema)
        df_ica_tot = run_query(QUERIES["ica_total_lifetime"],schema)
        df_summ    = run_query(QUERIES["ai_summaries_total"],schema)
        df_act     = run_query(QUERIES["platform_actions_summary"],schema)

    auto_cases,manual_cases = ica_counts(df_ica_tot)

    rows=[]
    if not df_frt.empty and len(df_frt)>=2:
        ff=safe_float(df_frt,"avg_frt_hours",idx=0); lf=safe_float(df_frt,"avg_frt_hours",idx=-2)
        rows.append({"Metric":"First response time","Baseline":f"{ff} hrs ({df_frt['month'].iloc[0]})",
                     "Current":f"{lf} hrs ({df_frt['month'].iloc[-2]})","Change":f"↓ {round((1-lf/ff)*100) if ff else 0}%",
                     "Agent":"Routing Agent","Status":"✅ Strong"})
    if not df_esc.empty:
        le=safe_float(df_esc,"escalation_pct",idx=-2)
        rows.append({"Metric":"Escalation rate","Baseline":f"{safe_float(df_esc,'escalation_pct',idx=0)}% ({df_esc['month'].iloc[0]})",
                     "Current":f"{le}% ({df_esc['month'].iloc[-2]})","Change":"✓ Below benchmark" if le<benchmark else "⚠ Above benchmark",
                     "Agent":"Escalation Agent","Status":"✅ On track" if le<benchmark else "⚠️ Watch"})
    if not df_lte.empty:
        prev=int(df_lte["cases_predicted"].sum())-int(df_lte["actually_escalated"].sum())
        rows.append({"Metric":"Escalations prevented (LTE)","Baseline":"—","Current":f"{prev:,} intercepted",
                     "Change":f"~{prev:,} avoided","Agent":"Escalation Agent","Status":"✅ Strong"})
    if not df_re.empty and len(df_re)>=2:
        fr=safe_float(df_re,"reassignment_pct",idx=0); lr=safe_float(df_re,"reassignment_pct",idx=-2)
        rows.append({"Metric":"Reassignment rate","Baseline":f"{fr}% ({df_re['month'].iloc[0]})",
                     "Current":f"{lr}% ({df_re['month'].iloc[-2]})","Change":f"↓ {round(fr-lr,1)}pp",
                     "Agent":"Routing Agent (ICA)","Status":"✅ Strong"})
    if ica_active(df_ica_tot):
        total=auto_cases+manual_cases
        rows.append({"Metric":"ICA auto-assignments","Baseline":"0 (at go-live)","Current":f"{auto_cases:,} lifetime",
                     "Change":f"{round(auto_cases/total*100) if total else 0}% auto rate","Agent":"Routing Agent (ICA)","Status":"✅ Active"})
    if not df_sent.empty:
        rows.append({"Metric":"Avg sentiment","Baseline":f"{safe_float(df_sent,'avg_sentiment',idx=0)} ({df_sent['month'].iloc[0]})",
                     "Current":f"{safe_float(df_sent,'avg_sentiment',idx=-1)}/100","Change":"Stable","Agent":"Sentiment Agent","Status":"✅ Stable"})
    if not df_alerts.empty:
        rows.append({"Metric":"Alerts engagement","Baseline":"—","Current":f"{int(df_alerts['total_alerts'].sum()):,} alerts",
                     "Change":f"Avg {round(df_alerts['total_alerts'].mean()):,}/month","Agent":"Escalation Agent","Status":"✅ Active"})
    if not df_summ.empty:
        rows.append({"Metric":"AI summaries generated","Baseline":"0 (at go-live)","Current":f"{int(df_summ['total'].sum()):,} total",
                     "Change":"Growing monthly","Agent":"Summarization Agent","Status":"✅ Growing"})

    st.subheader("What has improved since go-live")
    if rows:
        st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)

    st.divider()
    st.subheader("Visual highlights")
    col1,col2,col3=st.columns(3)
    if not df_frt.empty:
        with col1:
            st.markdown("**FRT journey since go-live**")
            st.plotly_chart(area_chart(df_frt,x="month",
                y_cols=[{"col":"avg_frt_hours","name":"FRT (hrs)","color":COLORS["blue"],"fill":"rgba(24,95,165,0.1)"}],
                height=200),use_container_width=True)
    if ica_active(df_ica_tot):
        total=auto_cases+manual_cases
        with col2:
            st.markdown("**ICA auto vs manual (lifetime)**")
            st.plotly_chart(donut_chart(["Auto","Manual"],[auto_cases,manual_cases],
                [COLORS["teal"],COLORS["gray"]],height=200,
                center_text=f"{round(auto_cases/total*100) if total else 0}% auto"),use_container_width=True)
    if not df_health.empty:
        with col3:
            st.markdown("**Account health score**")
            st.plotly_chart(gauge_chart(safe_float(df_health,"avg_health_score",idx=-1),label="/ 100",height=200),use_container_width=True)

    st.divider()
    st.subheader("⚠️ Watch items for renewal conversation")
    watch=[]
    if not df_health.empty:
        lh=safe_float(df_health,"avg_health_score",idx=-1); peak=float(df_health["avg_health_score"].max())
        if lh<peak-5: watch.append(f"**Account health score** dropped from {peak} (peak) to {lh}. Investigate at sub-account level.")
    if not df_esc.empty:
        le=safe_float(df_esc,"escalation_pct",idx=-2)
        if le>=benchmark: watch.append(f"**Escalation rate** ({le}%) at or above {benchmark}% benchmark.")
    if not df_act.empty and int(df_act["total_actions"].sum())<50:
        watch.append(f"**Platform engagement low** — only {int(df_act['total_actions'].sum())} agent actions recorded.")
    if watch:
        for w in watch: st.warning(w,icon="⚠️")
    else:
        st.success("No critical watch items — account metrics are in healthy range.",icon="✅")

    if customer.get("notes"):
        st.divider()
        st.markdown(f"**Account notes:** {customer['notes']}")

    st.divider()
    st.subheader("📥 Generate renewal deck")
    col_btn,col_info=st.columns([1,3])
    with col_btn:
        generate=st.button("🎯 Generate renewal deck",type="primary",use_container_width=True)
    with col_info:
        st.caption("Takes 15–30s · Calls Claude API once for narrative · Downloads as .pptx")

    if generate:
        from utils.slides_generator import generate_slides_deck, generate_slides_deck_with_charts
        deck_data = {
            "frt_monthly":          df_frt,
            "escalation_monthly":   df_esc,
            "lte_accuracy_monthly": df_lte,
            "reassignment_monthly": df_re,
            "account_health_monthly": df_health,
            "sentiment_monthly":    df_sent,
            "alerts_monthly":       df_alerts,
            "ica_total_lifetime":   df_ica_tot,
            "ai_summaries_total":   df_summ,
        }
        with st.spinner("Writing narrative · Building slides in Google Slides..."):
            try:
                pendo_data = None
                try:
                    from utils.pendo_conn import (get_page_views, get_feature_events,
                        get_page_map, get_feature_map, get_visitor_count,
                        dau, top_modules, top_features)
                    import pandas as _pd
                    pendo_id = customer.get("pendo_id", customer_name.lower())
                    _pm = get_page_map(); _fm = get_feature_map()
                    _dp = get_page_views(pendo_id, 30)
                    _df = get_feature_events(pendo_id, 30)
                    _dau_df = dau(_dp)
                    _pu = _dp.groupby("visitorId")["numMinutes"].sum().reset_index() if not _dp.empty else _pd.DataFrame()
                    _tot = len(_pu)
                    _a  = len(_pu[_pu["numMinutes"]>=180]) if not _pu.empty else 0
                    _mo = len(_pu[(_pu["numMinutes"]>=60)&(_pu["numMinutes"]<180)]) if not _pu.empty else 0
                    _lo = len(_pu[_pu["numMinutes"]<60]) if not _pu.empty else 0
                    _p  = lambda n: round(n/_tot*100) if _tot else 0
                    pendo_data = {
                        "visitors":     get_visitor_count(pendo_id),
                        "avg_dau":      round(float(_dau_df["dau"].mean())) if not _dau_df.empty else 0,
                        "days":         30,
                        "tiers":        {"active":_a,"moderate":_mo,"low":_lo,
                                         "active_pct":_p(_a),"moderate_pct":_p(_mo),
                                         "low_pct":_p(_lo),"total":_tot},
                        "top_modules":  top_modules(_dp, _pm, n=3),
                        "top_features": top_features(_df, _fm, n=8),
                    }
                except Exception:
                    pass
                url, charts_added = generate_slides_deck_with_charts(
                customer_name, customer, deck_data,
                pendo_data=pendo_data, include_charts=True)
                charts_info = f" + {len(charts_added)} chart slides" if charts_added else ""
                st.success(f"✅ Deck created in your Google Drive!{charts_info}")
                st.link_button("📊 Open in Google Slides →", url, type="primary")
                st.caption(f"Direct link: {url}")
            except Exception as e:
                st.error(f"Error: {e}", icon="❌")

# ════════════════════════════════════════════════════════════════════════════
# TAB 6 — PENDO ADOPTION
# ════════════════════════════════════════════════════════════════════════════
elif tab == 7:
    from utils.pendo_conn import (get_visitor_count, get_page_views, get_feature_events,
                                   get_page_map, get_feature_map,
                                   dau, wau, top_pages, top_features, time_in_platform,
                                   adoption_tiers, top_modules,
                                   geo_breakdown, region_summary, country_breakdown)
    from utils.snowflake_conn import run_query
    from utils.charts import COLORS, area_chart, bar_chart, line_chart, donut_chart
    from queries.registry import QUERIES

    pendo_id  = customer.get("pendo_id", customer_name.lower())
    pendo_ids = customer.get("pendo_ids", [pendo_id])
    col_d, _ = st.columns([1, 4])
    days = col_d.selectbox("Date range", [7, 14, 30, 60, 90], index=2,
                           format_func=lambda x: f"Last {x} days")

    with st.spinner("Loading Pendo data..."):
        import pandas as _pd2
        page_map    = get_page_map()
        feature_map = get_feature_map()
        visitors    = sum(get_visitor_count(p) for p in pendo_ids)
        _pf = [get_page_views(p, days) for p in pendo_ids]
        _ff = [get_feature_events(p, days) for p in pendo_ids]
        df_pages    = _pd2.concat([x for x in _pf if not x.empty], ignore_index=True).drop_duplicates() if any(not x.empty for x in _pf) else _pd2.DataFrame()
        df_features = _pd2.concat([x for x in _ff if not x.empty], ignore_index=True).drop_duplicates() if any(not x.empty for x in _ff) else _pd2.DataFrame()

    if df_pages.empty and df_features.empty:
        st.warning(f"No Pendo data for `{", ".join(pendo_ids)}` in last {days} days.", icon="⚠️")
        st.stop()

    # ── Derived metrics ───────────────────────────────────────────────────────
    df_dau   = dau(df_pages)
    df_wau   = wau(df_pages)
    df_time  = time_in_platform(df_pages)
    df_top_p = top_pages(df_pages, page_map)
    df_top_f = top_features(df_features, feature_map)
    tiers    = adoption_tiers(df_pages)
    df_mods  = top_modules(df_pages, page_map, n=3)
    df_geo   = geo_breakdown(df_pages)
    df_reg   = region_summary(df_geo)
    df_cntry = country_breakdown(df_geo, top_n=10)

    active_users  = int(df_pages["visitorId"].nunique()) if not df_pages.empty else 0
    total_views   = int(df_pages["numEvents"].sum())    if not df_pages.empty else 0
    total_clicks  = int(df_features["numEvents"].sum()) if not df_features.empty else 0
    avg_dau_val   = round(df_dau["dau"].mean())         if not df_dau.empty else 0

    # ── KPI row ───────────────────────────────────────────────────────────────
    c1,c2,c3,c4,c5 = st.columns(5)
    c1.metric("Total visitors (lifetime)", f"{visitors:,}")
    c2.metric(f"Active users (last {days}d)", f"{active_users:,}")
    c3.metric("Avg daily active users", f"{avg_dau_val:,}")
    c4.metric("Total page views", f"{total_views:,}")
    c5.metric("Feature clicks", f"{total_clicks:,}")


    # ── SL value insight banner ───────────────────────────────────────────────
    with st.spinner("Loading SL impact metrics..."):
        df_summ_ins  = run_query(QUERIES["account_summaries_insight"], schema)
        df_risk_ins  = run_query(QUERIES["at_risk_accounts_insight"],  schema)

    accounts_summarized = int(df_summ_ins["accounts_summarized"].iloc[0]) if not df_summ_ins.empty else 0
    at_risk             = int(df_risk_ins["at_risk_accounts"].iloc[0])    if not df_risk_ins.empty else 0
    total_scored        = int(df_risk_ins["total_accounts_scored"].iloc[0]) if not df_risk_ins.empty else 0

    # Hrs estimate — configurable assumption shown transparently
    mins_per_review = 20  # industry benchmark: ~20 min per manual account review
    weeks_in_period = 4   # monthly cadence assumption
    hours_saved = round((accounts_summarized * mins_per_review) / 60 / weeks_in_period, 1)

    col_ins1, col_ins2 = st.columns(2)
    with col_ins1:
        st.markdown(f"""<div style='background:linear-gradient(135deg,#0F2A1A,#0A1931);
            border:0.5px solid #1D9E75;border-radius:10px;padding:16px 20px'>
          <div style='font-size:11px;color:#1D9E75;font-weight:600;letter-spacing:.05em;
               text-transform:uppercase;margin-bottom:8px'>AI summarization impact</div>
          <div style='font-size:28px;font-weight:700;color:#e6edf3;margin-bottom:4px'>
            {accounts_summarized:,} accounts auto-summarized
          </div>
          <div style='font-size:13px;color:#8b949e;line-height:1.5'>
            SupportLogic AI generates account summaries automatically, saving managers an estimated
            <span style='color:#1D9E75;font-weight:600'>~{hours_saved} hrs/week</span>
            of manual review
          </div>
          <div style='font-size:10px;color:#555;margin-top:6px'>
            Estimate based on ~{mins_per_review} min/account (industry benchmark) on a monthly review cadence
          </div>
        </div>""", unsafe_allow_html=True)

    with col_ins2:
        risk_pct = round(at_risk / total_scored * 100) if total_scored else 0
        # Load at-risk account names
        df_risk_detail = run_query("""
            WITH latest AS (
                SELECT s_object_id_creator AS account_id, score_value,
                    ROW_NUMBER() OVER (PARTITION BY s_object_id_creator ORDER BY s_created_at DESC) AS rn
                FROM PIPE_DATABASE.<SCHEMA>.ml_prediction
                WHERE ml_prediction_type = 'ACCOUNT_HEALTH_SCORE'
            )
            SELECT account_id, ROUND(score_value,1) AS score
            FROM latest
            WHERE rn = 1 AND score_value < 60
            ORDER BY score_value ASC
            LIMIT 5
        """, schema)

        risk_names = ""
        if not df_risk_detail.empty:
            items = []
            for _, r in df_risk_detail.iterrows():
                items.append(f"<li style='margin:3px 0;color:#c9d1d9'>{r['account_id']} "
                             f"<span style='color:#E24B4A;font-size:11px'>({r['score']}/100)</span></li>")
            risk_names = "<ul style='margin:8px 0 0;padding-left:16px;font-size:12px'>" + "".join(items) + "</ul>"

        if total_scored <= 1:
            risk_note = "<div style='font-size:11px;color:#555;margin-top:6px'>Note: ML health scoring may be running at the case level rather than sub-account level for this customer — check ML_PREDICTION data granularity.</div>"
        else:
            risk_note = f"<div style='font-size:10px;color:#555;margin-top:6px'>Health score threshold: &lt;60/100 · {risk_pct}% of {total_scored:,} scored accounts</div>"

        st.markdown(f"""<div style='background:linear-gradient(135deg,#2A1010,#1A0A1A);
            border:0.5px solid #E24B4A;border-radius:10px;padding:16px 20px'>
          <div style='font-size:11px;color:#E24B4A;font-weight:600;letter-spacing:.05em;
               text-transform:uppercase;margin-bottom:8px'>Early risk detection</div>
          <div style='font-size:28px;font-weight:700;color:#e6edf3;margin-bottom:4px'>
            {at_risk:,} at-risk {"account" if at_risk == 1 else "accounts"}
          </div>
          <div style='font-size:13px;color:#8b949e;line-height:1.5'>
            surfaced early —
            <span style='color:#E24B4A;font-weight:600'>ahead of escalation or renewal risk</span>
          </div>
          {risk_names}
          {risk_note}
        </div>""", unsafe_allow_html=True)

    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    st.divider()

    # ── Section 1: User adoption tiers ────────────────────────────────────────
    st.subheader("👥 User adoption")
    st.caption(f"Based on time spent in platform over last {days} days · "
               f"Active = 3+ hrs · Moderate = 1–3 hrs · Low = <1 hr")

    tier_col1, tier_col2, tier_col3, tier_col4 = st.columns([1,1,1,2])

    with tier_col1:
        st.markdown(f"""<div style='background:#1a3a2a;border:0.5px solid #1D9E75;border-radius:10px;padding:14px 16px;text-align:center'>
          <div style='font-size:11px;color:#8b949e;margin-bottom:4px'>Active (3+ hrs)</div>
          <div style='font-size:36px;font-weight:600;color:#1D9E75'>{tiers["active_pct"]}%</div>
          <div style='font-size:12px;color:#8b949e'>{tiers["active"]} of {tiers["total"]} users</div>
        </div>""", unsafe_allow_html=True)

    with tier_col2:
        st.markdown(f"""<div style='background:#3a2e10;border:0.5px solid #BA7517;border-radius:10px;padding:14px 16px;text-align:center'>
          <div style='font-size:11px;color:#8b949e;margin-bottom:4px'>Moderate (1–3 hrs)</div>
          <div style='font-size:36px;font-weight:600;color:#BA7517'>{tiers["moderate_pct"]}%</div>
          <div style='font-size:12px;color:#8b949e'>{tiers["moderate"]} of {tiers["total"]} users</div>
        </div>""", unsafe_allow_html=True)

    with tier_col3:
        st.markdown(f"""<div style='background:#1e1e2e;border:0.5px solid #534AB7;border-radius:10px;padding:14px 16px;text-align:center'>
          <div style='font-size:11px;color:#8b949e;margin-bottom:4px'>Low (&lt;1 hr)</div>
          <div style='font-size:36px;font-weight:600;color:#534AB7'>{tiers["low_pct"]}%</div>
          <div style='font-size:12px;color:#8b949e'>{tiers["low"]} of {tiers["total"]} users</div>
        </div>""", unsafe_allow_html=True)

    with tier_col4:
        if not tiers["per_user"].empty:
            tier_color = {"Active": COLORS["green"], "Moderate": COLORS["amber"], "Low": COLORS["purple"]}
            df_t = tiers["per_user"].copy()
            df_t["color"] = df_t["tier"].map(tier_color)
            df_t["short_id"] = df_t["visitorId"].str.split("@").str[0]
            st.plotly_chart(bar_chart(
                df_t, x="short_id",
                y_cols=[{"col":"total_minutes","name":"Minutes","color":df_t["color"].tolist(),"colors":df_t["color"].tolist()}],
                height=180,
            ), use_container_width=True)

    st.divider()

    # ── Section 2: Top 3 modules ──────────────────────────────────────────────
    st.subheader("🏆 Top 3 modules")
    if not df_mods.empty:
        mod_cols = st.columns(3)
        rank_icons = ["🥇","🥈","🥉"]
        rank_colors = [COLORS["amber"], COLORS["gray"], COLORS["blue"]]
        for i, row in df_mods.iterrows():
            with mod_cols[i]:
                st.markdown(f"""<div style='background:#161b22;border:0.5px solid #30363d;border-radius:10px;padding:14px 16px'>
                  <div style='font-size:20px;margin-bottom:6px'>{rank_icons[i]}</div>
                  <div style='font-size:13px;font-weight:500;color:#e6edf3;margin-bottom:8px'>{row["page_name"]}</div>
                  <div style='display:flex;justify-content:space-between;margin-bottom:4px'>
                    <span style='font-size:11px;color:#8b949e'>Page views</span>
                    <span style='font-size:12px;font-weight:500;color:{rank_colors[i]}'>{int(row["total_views"]):,}</span>
                  </div>
                  <div style='display:flex;justify-content:space-between;margin-bottom:4px'>
                    <span style='font-size:11px;color:#8b949e'>Unique users</span>
                    <span style='font-size:12px;font-weight:500;color:#c9d1d9'>{int(row["unique_users"])}</span>
                  </div>
                  <div style='display:flex;justify-content:space-between'>
                    <span style='font-size:11px;color:#8b949e'>Time spent</span>
                    <span style='font-size:12px;font-weight:500;color:#c9d1d9'>{row["total_hrs"]} hrs</span>
                  </div>
                </div>""", unsafe_allow_html=True)
    else:
        st.info("No module data available.", icon="ℹ️")

    st.divider()

    # ── Section 3: Top 3 alerts YTD (from Snowflake) ─────────────────────────
    st.subheader("🔔 Top 3 alerts fired YTD")
    st.caption("Source: SupportLogic alert engine · Jan 2026 – present")

    with st.spinner("Loading alert data..."):
        df_alert_cols = run_query("""
            SELECT column_name FROM PIPE_DATABASE.information_schema.columns
            WHERE table_schema = '<SCHEMA>' AND table_name = 'ALERTS'
            ORDER BY ordinal_position
        """, schema)

    # Determine correct alert name/type column
    alert_name_col = None
    if not df_alert_cols.empty:
        cols_lower = df_alert_cols["column_name"].str.lower().tolist()
        for candidate in ["alert_type","alert_name","type","name","alert_definition_name","category"]:
            if candidate in cols_lower:
                alert_name_col = candidate.upper()
                break

    if alert_name_col:
        with st.spinner("Fetching top alerts..."):
            df_alerts_ytd = run_query(f"""
                SELECT {alert_name_col} AS alert_type,
                    COUNT(*) AS fired_count,
                    COUNT(DISTINCT sl_case_id) AS unique_cases
                FROM PIPE_DATABASE.<SCHEMA>.alerts
                WHERE s_created_at >= '2026-01-01'
                GROUP BY 1
                ORDER BY fired_count DESC
                LIMIT 3
            """, schema)

        if not df_alerts_ytd.empty:
            alert_cols = st.columns(3)
            alert_icons = ["🔴","🟠","🟡"]
            for i, row in df_alerts_ytd.iterrows():
                with alert_cols[i]:
                    st.markdown(f"""<div style='background:#161b22;border:0.5px solid #E24B4A;border-radius:10px;padding:14px 16px'>
                      <div style='font-size:18px;margin-bottom:6px'>{alert_icons[i]}</div>
                      <div style='font-size:12px;font-weight:500;color:#e6edf3;margin-bottom:8px'>{row["alert_type"]}</div>
                      <div style='display:flex;justify-content:space-between;margin-bottom:4px'>
                        <span style='font-size:11px;color:#8b949e'>Times fired</span>
                        <span style='font-size:12px;font-weight:500;color:#E24B4A'>{int(row["fired_count"]):,}</span>
                      </div>
                      <div style='display:flex;justify-content:space-between'>
                        <span style='font-size:11px;color:#8b949e'>Unique cases</span>
                        <span style='font-size:12px;font-weight:500;color:#c9d1d9'>{int(row["unique_cases"]):,}</span>
                      </div>
                    </div>""", unsafe_allow_html=True)
        else:
            st.info("No alert data found for this customer YTD.", icon="ℹ️")
    else:
        # Fallback: show alert counts without type breakdown
        with st.spinner("Fetching alert summary..."):
            df_alert_sum = run_query("""
                SELECT COUNT(*) AS total_alerts,
                    COUNT(DISTINCT sl_case_id) AS unique_cases,
                    MIN(s_created_at) AS first_alert,
                    MAX(s_created_at) AS last_alert
                FROM PIPE_DATABASE.<SCHEMA>.alerts
                WHERE s_created_at >= '2026-01-01'
            """, schema)
        if not df_alert_sum.empty:
            row = df_alert_sum.iloc[0]
            st.metric("Total alerts fired YTD", f"{int(row['total_alerts']):,}",
                      f"Across {int(row['unique_cases']):,} unique cases")
        st.caption("Alert type breakdown not available — column schema differs for this customer.")

    st.divider()

    # ── Section: Regional breakdown ──────────────────────────────────────────
    if not df_reg.empty:
        st.divider()
        st.subheader("🌍 Regional user distribution")
        st.caption(f"Based on geolocation of Pendo page events — last {days} days · "
                   "Region = user's most frequent login location")

        # Region cards
        region_colors = {
            "Americas": ("#185FA5", "#0A1931", "🌎"),
            "EMEA":     ("#0F6E56", "#0A1931", "🌍"),
            "APAC":     ("#534AB7", "#0A1931", "🌏"),
            "Unknown":  ("#B4B2A9", "#161b22", "❓"),
        }
        reg_cols = st.columns(len(df_reg))
        for i, row in df_reg.iterrows():
            reg = row["geo_region"]
            color, bg, icon = region_colors.get(reg, ("#B4B2A9","#161b22","🌐"))
            with reg_cols[i]:
                st.markdown(f"""<div style='background:{bg};border:0.5px solid {color};
                    border-radius:10px;padding:14px 16px;text-align:center'>
                  <div style='font-size:22px;margin-bottom:4px'>{icon}</div>
                  <div style='font-size:11px;color:#8b949e;margin-bottom:4px'>{reg}</div>
                  <div style='font-size:28px;font-weight:700;color:{color}'>{row["users"]}</div>
                  <div style='font-size:11px;color:#8b949e'>users ({row["pct"]}%)</div>
                  <div style='font-size:10px;color:#555;margin-top:4px'>{row["sessions"]:,} sessions</div>
                </div>""", unsafe_allow_html=True)

        # Country breakdown
        if not df_cntry.empty:
            st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
            st.markdown("**Top countries by active users**")
            col_bar, col_tbl = st.columns([2, 1])
            with col_bar:
                from utils.charts import COLORS, bar_chart
                clrs = [COLORS["blue"],COLORS["teal"],COLORS["purple"],COLORS["amber"],
                        COLORS["red"],COLORS["green"],COLORS["coral"],COLORS["gray"],
                        COLORS["blue"],COLORS["teal"]]
                st.plotly_chart(bar_chart(
                    df_cntry.sort_values("users"), x="country",
                    y_cols=[{"col":"users","name":"Users",
                             "color":[clrs[i%len(clrs)] for i in range(len(df_cntry))],
                             "colors":[clrs[i%len(clrs)] for i in range(len(df_cntry))]}],
                    height=280, horizontal=True), use_container_width=True)
            with col_tbl:
                df_cntry_display = df_cntry.copy()
                df_cntry_display.columns = ["Country", "Users"]
                df_cntry_display["% of total"] = (
                    df_cntry_display["Users"] / df_cntry_display["Users"].sum() * 100
                ).round(1).astype(str) + "%"
                st.dataframe(df_cntry_display, use_container_width=True,
                             hide_index=True)

    st.divider()
        # ── Section 4: DAU trend + time in platform ───────────────────────────────
    col_l, col_r = st.columns(2)
    if not df_dau.empty:
        with col_l:
            st.markdown("**Daily active users (DAU)**")
            st.plotly_chart(area_chart(df_dau, x="date",
                y_cols=[{"col":"dau","name":"Active users","color":COLORS["blue"],"fill":"rgba(24,95,165,0.1)"}],
                height=220), use_container_width=True)
    if not df_time.empty:
        with col_r:
            st.markdown("**Time in platform — daily (minutes)**")
            st.plotly_chart(area_chart(df_time, x="date",
                y_cols=[{"col":"total_minutes","name":"Minutes","color":COLORS["teal"],"fill":"rgba(15,110,86,0.1)"}],
                height=220), use_container_width=True)

    # ── Section 5: Top pages + features ──────────────────────────────────────
    col_l2, col_r2 = st.columns(2)
    if not df_top_p.empty:
        with col_l2:
            st.markdown(f"**Top pages — last {days} days**")
            st.plotly_chart(bar_chart(
                df_top_p.sort_values("total_views"), x="page_name",
                y_cols=[{"col":"total_views","name":"Views","color":COLORS["blue"]}],
                height=300, horizontal=True), use_container_width=True)
    if not df_top_f.empty:
        with col_r2:
            palette = [COLORS["teal"],COLORS["blue"],COLORS["purple"],COLORS["amber"],
                       COLORS["red"],COLORS["green"],COLORS["coral"],COLORS["gray"]]
            clrs = [palette[i%len(palette)] for i in range(len(df_top_f))]
            st.markdown(f"**Top feature clicks — last {days} days**")
            st.plotly_chart(bar_chart(
                df_top_f.sort_values("total_clicks"), x="feature_name",
                y_cols=[{"col":"total_clicks","name":"Clicks","color":clrs,"colors":clrs}],
                height=300, horizontal=True), use_container_width=True)

    st.caption(f"Pendo account: `{pendo_id}` · Alerts: PIPE_DATABASE.{schema}.ALERTS · "
               f"Pages: {len(page_map):,} · Features: {len(feature_map):,}")

elif tab == 8:
    st.markdown("""
<div style='text-align:center;padding:80px 20px'>
  <div style='font-size:48px;margin-bottom:16px'>🔧</div>
  <div style='font-size:20px;font-weight:600;color:#e6edf3;margin-bottom:8px'>Under Construction</div>
  <div style='font-size:13px;color:#8b949e;line-height:1.8'>
    This section is reserved for technical use and will be available in a future update.
  </div>
</div>""", unsafe_allow_html=True)
    st.stop()

    from utils.snowflake_conn import run_query
    from queries.registry import QUERIES, QUERY_CATALOG

    st.subheader("⚙️ Query explorer")
    st.caption(f"Running against: `PIPE_DATABASE.{schema}` ({customer_name})")

    t1,t2=st.tabs(["Browse & run","How to add a metric"])
    with t1:
        agents=["All"]+sorted(set(q["agent"] for q in QUERY_CATALOG))
        tabs_f=["All"]+sorted(set(q["tab"] for q in QUERY_CATALOG))
        cf1,cf2=st.columns(2)
        fa=cf1.selectbox("Filter by agent",agents)
        ft=cf2.selectbox("Filter by tab",tabs_f)
        catalog=[q for q in QUERY_CATALOG if (fa=="All" or q["agent"]==fa) and (ft=="All" or q["tab"]==ft)]
        for entry in catalog:
            badge="🟢" if entry["status"]=="stable" else "🟡"
            with st.expander(f"{badge} **{entry['name']}** — {entry['agent']}"):
                st.markdown(f"**Tables:** `{'`, `'.join(entry['tables'])}`")
                st.code(QUERIES[entry["id"]].strip(),language="sql")
                if st.button("▶ Run",key=f"run_{entry['id']}"):
                    with st.spinner("Running..."):
                        df=run_query(QUERIES[entry["id"]],schema)
                    if not df.empty:
                        st.success(f"✓ {len(df)} rows")
                        st.dataframe(df,use_container_width=True)
                        st.download_button("⬇ CSV",df.to_csv(index=False),
                            f"{customer_name}_{entry['id']}.csv","text/csv",key=f"dl_{entry['id']}")
                    else:
                        st.warning("No rows returned.")
    with t2:
        st.markdown("""
### Adding a new metric — 3 steps
**Step 1** — Add SQL to `queries/registry.py` with `<SCHEMA>` placeholder
**Step 2** — Add entry to `QUERY_CATALOG` in the same file
**Step 3** — Add chart in the relevant tab section in `Home.py`

Available chart types: `line_chart`, `area_chart`, `bar_chart`, `combo_chart`, `stacked_bar`, `donut_chart`, `gauge_chart`, `bubble_chart`, `step_line`
""")

# ════════════════════════════════════════════════════════════════════════════
# TAB 8 — GCP LOG TROUBLESHOOTER
# Source: Data Pipeline Runbook (Confluence 1778089987)
#         ACA Logs & Debugging Guide (Confluence 2165932051)
#         Debugging ACA + ML Prediction (Confluence 1512767532)
# ════════════════════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════════════════════
# TAB 8 — GCP LOG TROUBLESHOOTER
# Source: Data Pipeline Runbook (Confluence 1778089987)
#         ACA Logs & Debugging Guide (Confluence 2165932051)
#         Debugging ACA + ML Prediction (Confluence 1512767532)
# ════════════════════════════════════════════════════════════════════════════

elif tab == 9:
    st.markdown("""
<div style='text-align:center;padding:80px 20px'>
  <div style='font-size:48px;margin-bottom:16px'>🔧</div>
  <div style='font-size:20px;font-weight:600;color:#e6edf3;margin-bottom:8px'>Under Construction</div>
  <div style='font-size:13px;color:#8b949e;line-height:1.8'>
    This section is reserved for technical use and will be available in a future update.
  </div>
</div>""", unsafe_allow_html=True)

    import urllib.parse
    from datetime import datetime, timedelta, timezone

    GCP_PROJECT = "supportlogic"

    # ── Customer hostname mapping ─────────────────────────────────────────────
    HOSTNAME_OVERRIDES = {}
    hostname = HOSTNAME_OVERRIDES.get(
        customer_name,
        customer.get("gcp_hostname", customer_name.lower().replace(" ", ""))
    )

    # ── Container catalogue from Data Pipeline Runbook ────────────────────────
    # layer: "backend" = customer-backend server
    #        "frontend" = customer UI server
    #        "shared"   = idt-backend (shared across all customers)
    #        "ml"       = ML service (shared)
    CONTAINERS = {
        # ── CRM Ingestion ──────────────────────────────────────────────────
        "crm_importer": {
            "layer": "backend",
            "desc": "Fetches CRM data → uds_case, uds_comment, uds_user, uds_account, sl_case_history",
            "pipeline_stage": "CRM ingestion",
            "symptom": "Missing cases, comments, or users in SL UI",
        },
        # ── Data processing pipeline ───────────────────────────────────────
        "comment_metrics": {
            "layer": "backend",
            "desc": "uds_comment → sl_comment_metric_2020_02",
            "pipeline_stage": "Metrics",
            "symptom": "Missing comment metrics",
        },
        "case_progress": {
            "layer": "backend",
            "desc": "sl_case_history → sl_case_progress",
            "pipeline_stage": "Case progress",
            "symptom": "Incorrect case progress data",
        },
        "user_metrics": {
            "layer": "backend",
            "desc": "uds_case/comment/user → sl_user_metrics",
            "pipeline_stage": "User metrics",
            "symptom": "Incorrect agent metrics",
        },
        "user_stitcher": {
            "layer": "backend",
            "desc": "uds_user + sl_user_metrics → user",
            "pipeline_stage": "User stitching",
            "symptom": "User data missing or stale",
        },
        # ── ML / NLP pipeline ──────────────────────────────────────────────
        "email_predictor": {
            "layer": "ml",
            "desc": "uds_comment → span_doc_email (ML)",
            "pipeline_stage": "NLP — email classification",
            "symptom": "No sentiment signals, missing spans",
        },
        "diverse_v3": {
            "layer": "ml",
            "desc": "span_doc_email → span_doc_diverse (ML)",
            "pipeline_stage": "NLP — diverse signals",
            "symptom": "Missing sentiment categories",
        },
        "flex_ner": {
            "layer": "ml",
            "desc": "ontology + uds_case + span_doc_email → span_doc_ner (ML)",
            "pipeline_stage": "NLP — named entity recognition",
            "symptom": "Missing entity/keyword detection",
        },
        "span_stitcher": {
            "layer": "backend",
            "desc": "span_doc_diverse + email + sl_feedback → span_doc_consolidated",
            "pipeline_stage": "Span stitching",
            "symptom": "Sentiments missing or very few in dashboard",
        },
        "impulse_scorer": {
            "layer": "backend",
            "desc": "span_doc_consolidated → sl_impulse_score_by_channel_2020_02",
            "pipeline_stage": "Sentiment scoring",
            "symptom": "No sentiment scores visible",
        },
        "customer_scorer": {
            "layer": "backend",
            "desc": "sl_impulse_score_summary → sl_customer_score",
            "pipeline_stage": "Account scoring",
            "symptom": "Account health scores stale",
        },
        "feedback_sweeper": {
            "layer": "backend",
            "desc": "sl_feedback (RDB) → sl_feedback (PIPE SQL)",
            "pipeline_stage": "Feedback sync",
            "symptom": "Agent signal feedback not reflected",
        },
        "escalation_predictor": {
            "layer": "ml",
            "desc": "ML service for escalation prediction",
            "pipeline_stage": "LTE / Escalation prediction",
            "symptom": "LTE predictions not updating",
        },
        "escalation_prediction_summarizer": {
            "layer": "backend",
            "desc": "escalation_predictions → escalation_predictions_summary",
            "pipeline_stage": "Escalation summarization",
            "symptom": "Escalation summary stale",
        },
        # ── Case summary & alerting ────────────────────────────────────────
        "account_stitcher": {
            "layer": "backend",
            "desc": "uds_account + sl_customer_score → account",
            "pipeline_stage": "Account stitching",
            "symptom": "Account data stale or missing",
        },
        "entity_stitcher": {
            "layer": "backend",
            "desc": "span_doc_ner → entity_row, entity_summary",
            "pipeline_stage": "Entity stitching",
            "symptom": "Entities/keywords missing",
        },
        "case_summarizer": {
            "layer": "backend",
            "desc": "uds_case + all metrics → case_summary",
            "pipeline_stage": "Case summarization",
            "symptom": "Case summary data stale",
        },
        "case_summary_alerter": {
            "layer": "backend",
            "desc": "case_summary → alert_events",
            "pipeline_stage": "Alert triggering",
            "symptom": "Alerts not firing on case changes",
        },
        "timed_alerter": {
            "layer": "backend",
            "desc": "case_summary → alert_events (time-based)",
            "pipeline_stage": "Timed alerts",
            "symptom": "Time-based alerts not firing",
        },
        "alert_router": {
            "layer": "backend",
            "desc": "alert_events → alerts (routes to Slack/email/Teams)",
            "pipeline_stage": "Alert routing",
            "symptom": "Alerts fired but not delivered",
        },
        # ── Broker services (PIPE → UI RDB) ───────────────────────────────
        "case_summary_broker": {
            "layer": "frontend",
            "desc": "case_summary (PIPE) → case_summary (UI RDB). Missing cases in SL UI.",
            "pipeline_stage": "Broker — cases",
            "symptom": "Cases missing or stale in SL UI",
        },
        "unified_span_broker": {
            "layer": "frontend",
            "desc": "uds_comment + impulse_score → unified_spans (UI RDB)",
            "pipeline_stage": "Broker — spans/sentiments",
            "symptom": "Sentiments not showing in case view",
        },
        "account_broker": {
            "layer": "frontend",
            "desc": "account (PIPE) → companies (UI RDB)",
            "pipeline_stage": "Broker — accounts",
            "symptom": "Account data missing in SL UI",
        },
        "ontology_broker": {
            "layer": "frontend",
            "desc": "ontology_v4 → sl_ontology (UI RDB)",
            "pipeline_stage": "Broker — ontology",
            "symptom": "Signal categories not loading",
        },
        "user_broker": {
            "layer": "frontend",
            "desc": "user (PIPE) → users (UI RDB)",
            "pipeline_stage": "Broker — users",
            "symptom": "User/agent data stale in UI",
        },
        # ── Search / Elastic ───────────────────────────────────────────────
        "elastic_feeder": {
            "layer": "backend",
            "desc": "uds_case/comment/account/user → Elasticsearch",
            "pipeline_stage": "Search indexing",
            "symptom": "Search results missing or returning no results",
        },
        # ── ICA / Assignment ───────────────────────────────────────────────
        "aca_service": {
            "layer": "shared",
            "desc": "ICA daemon — auto-assignment of cases to agents",
            "pipeline_stage": "Case assignment (ICA)",
            "symptom": "Cases not being auto-assigned",
        },
        # ── UI / iframe ────────────────────────────────────────────────────
        "iframe_v2_ui": {
            "layer": "frontend",
            "desc": "SL embedded iframe widget served to CRM",
            "pipeline_stage": "UI",
            "symptom": "iframe not loading, widget errors, 500s",
        },
        # ── Writeback / UWF ────────────────────────────────────────────────
        "webhook_event_propagation_daemon": {
            "layer": "shared",
            "desc": "UWF — writes SL signals back to CRM (sentiment, health score, escalation flag)",
            "pipeline_stage": "CRM writeback (UWF)",
            "symptom": "CRM fields not updating (health score, sentiment etc)",
        },
        "dbss": {
            "layer": "frontend",
            "desc": "Database service — serves data to SL UI",
            "pipeline_stage": "UI data service",
            "symptom": "UI data not loading, API errors",
        },
        "backend_init": {
            "layer": "backend",
            "desc": "Bootstrap container — used to run static jobs and ack pubsub batches",
            "pipeline_stage": "Ops / maintenance",
            "symptom": "Used for manual interventions only",
        },
    }

    LAYER_COLOR = {
        "backend":  ("#1a3a2a", "#1D9E75", "🟢 Backend"),
        "frontend": ("#1a1a3a", "#534AB7", "🟣 Frontend"),
        "shared":   ("#3a2e10", "#BA7517", "🟡 Shared"),
        "ml":       ("#2a1a3a", "#E24B4A", "🔴 ML Service"),
    }

    # ── Sentiment pipeline ordered ─────────────────────────────────────────
    SENTIMENT_PIPELINE = [
        "email_predictor", "diverse_v3", "flex_ner",
        "span_stitcher", "impulse_scorer", "unified_span_broker"
    ]

    st.subheader("🔍 GCP Log Troubleshooter")
    st.caption(
        f"Customer: **{customer_name}** · Hostname: `{hostname}` · "
        f"Project: `{GCP_PROJECT}` · "
        f"Source: [Data Pipeline Runbook](https://supportlogic.atlassian.net/wiki/spaces/ENG/pages/1778089987)"
    )
    st.info(
        "Click **Open in Log Explorer** to open GCP Logs Explorer pre-filtered in your browser. "
        "Must be signed in with your SupportLogic Google account.",
        icon="ℹ️"
    )

    # ── Controls ──────────────────────────────────────────────────────────────
    col_t1, col_t2, col_t3, col_t4 = st.columns([1.2, 1, 1.2, 1.6])
    time_window = col_t1.selectbox("Time window",
        ["Last 1 hour", "Last 6 hours", "Last 24 hours", "Last 3 days", "Last 7 days"], index=1)
    severity = col_t2.selectbox("Min severity",
        ["DEFAULT", "INFO", "WARNING", "ERROR", "CRITICAL"], index=2)
    case_id = col_t3.text_input("Case ID", placeholder="e.g. 500Vy00000vUZ7fIAG")
    keyword = col_t4.text_input("Keyword / message filter", placeholder="e.g. CASE_SKIPPED or error text")

    window_map = {
        "Last 1 hour":   timedelta(hours=1),
        "Last 6 hours":  timedelta(hours=6),
        "Last 24 hours": timedelta(hours=24),
        "Last 3 days":   timedelta(days=3),
        "Last 7 days":   timedelta(days=7),
    }
    duration_map = {
        "Last 1 hour":   "PT1H",
        "Last 6 hours":  "PT6H",
        "Last 24 hours": "PT24H",
        "Last 3 days":   "P3D",
        "Last 7 days":   "P7D",
    }
    delta = window_map[time_window]
    now   = datetime.now(timezone.utc)
    start = now - delta

    def build_url(filter_lines: list) -> str:
        parts = list(filter_lines)
        if severity != "DEFAULT":
            parts.append(f"severity>={severity}")
        if case_id.strip():
            parts.append(f'SEARCH("{case_id.strip()}")')
        if keyword.strip():
            parts.append(f'jsonPayload.MESSAGE=~"{keyword.strip()}"')
        query   = "\n".join(parts)
        encoded = urllib.parse.quote(query, safe="")
        return (f"https://console.cloud.google.com/logs/query"
                f";query={encoded}"
                f";duration={duration_map[time_window]}"
                f"?project={GCP_PROJECT}")

    def filter_preview(filter_lines: list) -> str:
        parts = list(filter_lines)
        if severity != "DEFAULT":
            parts.append(f"severity>={severity}")
        if case_id.strip():
            parts.append(f'SEARCH("{case_id.strip()}")')
        if keyword.strip():
            parts.append(f'jsonPayload.MESSAGE=~"{keyword.strip()}"')
        return "\n".join(parts)

    def container_hostname(c_name: str) -> str:
        """Return the right _HOSTNAME pattern for a container."""
        layer = CONTAINERS.get(c_name, {}).get("layer", "backend")
        if layer == "shared" or layer == "ml":
            return "idt-backend"
        if layer == "backend":
            return f"{hostname}-backend"
        # frontend containers use just hostname (e.g. "fourth")
        return hostname

    def render_container_card(c_name: str, extra_filters: list = None):
        info = CONTAINERS.get(c_name, {})
        layer = info.get("layer", "backend")
        bg, border, layer_label = LAYER_COLOR.get(layer, LAYER_COLOR["backend"])
        h = container_hostname(c_name)
        base = [
            f'jsonPayload.CONTAINER_NAME="{c_name}"',
            f'jsonPayload._HOSTNAME=~"{h}"',
        ]
        if extra_filters:
            base.extend(extra_filters)
        url = build_url(base)

        col_info, col_btn = st.columns([3, 1])
        with col_info:
            st.markdown(
                f"<div style='background:{bg};border:0.5px solid {border};"
                f"border-radius:8px;padding:10px 14px;margin-bottom:4px'>"
                f"<div style='font-size:13px;font-weight:500;color:#e6edf3'>"
                f"<code style='background:transparent;color:{border}'>{c_name}</code>"
                f"  <span style='font-size:10px;color:#8b949e;margin-left:6px'>{layer_label}</span></div>"
                f"<div style='font-size:11px;color:#8b949e;margin-top:3px'>{info.get('desc','')}</div>"
                f"<div style='font-size:11px;color:#BA7517;margin-top:2px'>⚠ {info.get('symptom','')}</div>"
                f"</div>",
                unsafe_allow_html=True,
            )
            with st.expander("Show filter", expanded=False):
                st.code(filter_preview(base), language="text")
        with col_btn:
            st.link_button("Open in Log Explorer →", url, use_container_width=True)

    st.divider()

    # ── Legend ────────────────────────────────────────────────────────────────
    st.markdown("**Container layer legend:**")
    leg_cols = st.columns(4)
    for i, (layer, (bg, border, label)) in enumerate(LAYER_COLOR.items()):
        leg_cols[i].markdown(
            f"<div style='background:{bg};border:0.5px solid {border};border-radius:6px;"
            f"padding:6px 10px;font-size:12px;color:#e6edf3'>{label}</div>",
            unsafe_allow_html=True
        )
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    # ── Quick issue templates ─────────────────────────────────────────────────
    st.subheader("🚨 Quick issue templates")
    st.caption("Start here — each template maps to a known symptom")

    ISSUE_TEMPLATES = [
        {
            "title": "Cases / comments missing in SL UI",
            "desc": "Check CRM ingestion → case broker → search indexing chain",
            "containers": ["crm_importer", "case_summary_broker", "elastic_feeder"],
        },
        {
            "title": "Sentiments missing or very few",
            "desc": "Full NLP pipeline: email_predictor → diverse_v3 → flex_ner → span_stitcher → impulse_scorer → unified_span_broker",
            "containers": SENTIMENT_PIPELINE,
        },
        {
            "title": "Cases not auto-assigning (ICA)",
            "desc": "ICA daemon on shared backend — check for CASE_SKIPPED, ML failures, no agents in queue",
            "containers": ["aca_service"],
            "extra": {'aca_service': ['jsonPayload._HOSTNAME="idt-backend"']},
        },
        {
            "title": "Alerts not firing or not delivered",
            "desc": "Alert generation + routing chain",
            "containers": ["case_summary_alerter", "timed_alerter", "alert_router"],
        },
        {
            "title": "CRM writeback not working (UWF)",
            "desc": "Webhook propagation daemon — checks health score, sentiment, escalation flag writebacks",
            "containers": ["webhook_event_propagation_daemon"],
            "extra": {"webhook_event_propagation_daemon": [f'jsonPayload._HOSTNAME=~"{hostname}-backend"']},
        },
        {
            "title": "iframe widget errors / 500s",
            "desc": "Frontend iframe service for this customer",
            "containers": ["iframe_v2_ui", "dbss"],
        },
        {
            "title": "LTE / escalation predictions stale",
            "desc": "ML prediction service and summarizer",
            "containers": ["escalation_predictor", "escalation_prediction_summarizer"],
        },
        {
            "title": "Account / user data stale in UI",
            "desc": "Account and user stitching + broker services",
            "containers": ["account_stitcher", "user_stitcher", "account_broker", "user_broker"],
        },
    ]

    for tmpl in ISSUE_TEMPLATES:
        with st.expander(f"**{tmpl['title']}** — {tmpl['desc']}", expanded=False):
            extra_map = tmpl.get("extra", {})
            for c_name in tmpl["containers"]:
                info = CONTAINERS.get(c_name, {})
                layer = info.get("layer", "backend")
                h = extra_map.get(c_name, [f'jsonPayload._HOSTNAME=~"{container_hostname(c_name)}"'])
                base = [f'jsonPayload.CONTAINER_NAME="{c_name}"'] + h
                url  = build_url(base)
                _, border, layer_label = LAYER_COLOR.get(layer, LAYER_COLOR["backend"])
                col_i, col_b = st.columns([3, 1])
                with col_i:
                    st.markdown(
                        f"<code style='color:{border}'>{c_name}</code>"
                        f" <span style='font-size:10px;color:#8b949e'>{layer_label} · {info.get('pipeline_stage','')}</span>"
                        f"<br><span style='font-size:11px;color:#BA7517'>⚠ {info.get('symptom','')}</span>",
                        unsafe_allow_html=True,
                    )
                with col_b:
                    st.link_button("Open →", url, use_container_width=True)
                st.markdown("<div style='height:2px'></div>", unsafe_allow_html=True)

    st.divider()

    # ── All containers by layer ───────────────────────────────────────────────
    st.subheader("📦 All containers — browse by layer")

    layer_tabs = st.tabs(["🟢 Backend", "🟣 Frontend", "🟡 Shared", "🔴 ML Services"])
    layer_map  = {"backend": 0, "frontend": 1, "shared": 2, "ml": 3}

    for c_name, info in CONTAINERS.items():
        layer = info.get("layer", "backend")
        with layer_tabs[layer_map[layer]]:
            render_container_card(c_name)

    st.divider()

    # ── Custom query builder ──────────────────────────────────────────────────
    st.subheader("🔧 Custom query builder")
    col_c1, col_c2 = st.columns(2)
    custom_container = col_c1.text_input("CONTAINER_NAME", placeholder="e.g. span_stitcher")
    custom_hostname  = col_c2.text_input("_HOSTNAME", value=hostname)
    custom_message   = st.text_input("MESSAGE contains", placeholder="e.g. error or CASE_SKIPPED")
    custom_extra     = st.text_area("Additional filter lines (one per line)",
                                     placeholder='jsonPayload.MESSAGE=~"my_pattern"', height=80)

    if st.button("🔗 Build & open", type="primary"):
        lines = []
        if custom_container:
            lines.append(f'jsonPayload.CONTAINER_NAME="{custom_container}"')
        if custom_hostname:
            lines.append(f'jsonPayload._HOSTNAME=~"{custom_hostname}"')
        if custom_message:
            lines.append(f'jsonPayload.MESSAGE=~"{custom_message}"')
        if custom_extra.strip():
            lines.extend([l.strip() for l in custom_extra.strip().split("\n") if l.strip()])
        url = build_url(lines)
        st.code(filter_preview(lines), language="text")
        st.link_button("Open in Log Explorer →", url, type="primary")

    st.divider()
    st.caption(
        f"Hostname: `{hostname}` · Override via `gcp_hostname` in `utils/customers.py` · "
        f"[Data Pipeline Runbook](https://supportlogic.atlassian.net/wiki/spaces/ENG/pages/1778089987) · "
        f"[ACA Debugging Guide](https://supportlogic.atlassian.net/wiki/spaces/ENG/pages/2165932051)"
    )

# ════════════════════════════════════════════════════════════════════════════
# TAM DASHBOARD PAGE
# Paste this as elif tab == X in Home.py
# Add ("👤", "TAM") to TABS list
# ════════════════════════════════════════════════════════════════════════════


elif tab == 10:
    def try_arr(v):
        try: return f"${float(v)/1000:.0f}K"
        except: return "—"
    import pandas as pd
    import requests as _req
    from io import StringIO

    st.markdown("""
<div style='display:flex;align-items:center;gap:12px;margin-bottom:4px'>
  <div style='font-size:20px;font-weight:700;color:#e6edf3'>👤 TAM / TSE Account Dashboard</div>
</div>""", unsafe_allow_html=True)

    # ── Live Google Sheets fetch ───────────────────────────────────────────────
    SHEET_ID  = "1QhcklrfWoEyHRRmRIdzC0qUQ1VqiYkK3"
    SHEET_GID = "1187190574"  # "Up to date" tab

    @st.cache_data(ttl=300, show_spinner=False)  # refresh every 5 min
    def fetch_sheet(sheet_id, gid):
        """Fetch Google Sheet as CSV — works if sheet is shared or via OAuth."""
        # Try public CSV export first
        url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
        try:
            r = _req.get(url, timeout=15)
            if r.status_code == 200 and "DOCTYPE" not in r.text[:100]:
                return pd.read_csv(StringIO(r.text), dtype=str)
        except Exception:
            pass

        # Fallback: read from local file if available
        import pathlib as _pl
        local = _pl.Path("data/accounts.xlsx")
        if local.exists():
            import openpyxl as _xl
            wb = _xl.load_workbook(str(local), read_only=True, data_only=True)
            ws = wb["Up to date"]
            rows = list(ws.iter_rows(values_only=True))
            headers = [str(h).strip() if h else f"col_{i}" for i,h in enumerate(rows[0])]
            data = [dict(zip(headers, [str(v) if v is not None else "" for v in r]))
                    for r in rows[1:] if any(v is not None for v in r)]
            wb.close()
            return pd.DataFrame(data)
        return pd.DataFrame()

    @st.cache_data(ttl=300, show_spinner=False)
    def fetch_projects(sheet_id):
        url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid=0"
        sheets_to_try = ["Ongoing Projects", "Incoming"]
        # Fallback to local
        import pathlib as _pl
        local = _pl.Path("data/accounts.xlsx")
        result = {}
        if local.exists():
            import openpyxl as _xl
            wb = _xl.load_workbook(str(local), read_only=True, data_only=True)
            for sname in ["Ongoing Projects", "Incoming"]:
                if sname in wb.sheetnames:
                    ws = wb[sname]
                    rows = list(ws.iter_rows(values_only=True))
                    # skip first row (formula), use second as header
                    if len(rows) >= 2:
                        headers = [str(h).strip() if h else f"col_{i}"
                                   for i,h in enumerate(rows[1])]
                        data = [dict(zip(headers,[str(v) if v is not None else ""
                                                  for v in r]))
                                for r in rows[2:] if any(v is not None for v in r)]
                        result[sname] = pd.DataFrame(data)
            wb.close()
        return result

    col_refresh, col_info = st.columns([1, 4])
    with col_refresh:
        if st.button("🔄 Refresh data", use_container_width=True):
            st.cache_data.clear()
            st.rerun()
    with col_info:
        st.caption("Auto-refreshes every 5 min from Google Sheets · Manual refresh clears cache · "
                   "Falls back to local accounts.xlsx if sheet is not public")

    with st.spinner("Loading from Google Sheets..."):
        df = fetch_sheet(SHEET_ID, SHEET_GID)
        proj_data = fetch_projects(SHEET_ID)

    if df.empty:
        st.error("Could not load account data. Make sure the Google Sheet is shared publicly "
                 "OR accounts.xlsx is in the data/ folder.", icon="❌")
        st.info("To make the sheet public: Google Sheets → Share → Anyone with link → Viewer", icon="ℹ️")
        st.stop()

    # ── Column mapping ────────────────────────────────────────────────────────
    def col(df, *cands):
        for c in cands:
            if c in df.columns: return c
            m = [x for x in df.columns if c.lower() in x.lower()]
            if m: return m[0]
        return None

    C_CUST    = col(df, "Customer name", "Customer")
    C_TSE     = col(df, "Assigned TSE", "TSE")
    C_CSM     = col(df, "Customer Success Manager", "CSM")
    C_ARR     = col(df, "ARR")
    C_RENEW   = col(df, "Renewal Date")
    C_STAGE   = col(df, "Lifecycle Stage")
    C_CORE    = col(df, "Core Status")
    C_ASSIGN  = col(df, "Assign Status", "Assign purchased")
    C_ASSIST  = col(df, "Assist Status", "Assist purchased")
    C_EXPAND  = col(df, "Expand Status", "Expand purchased")
    C_ELEVATE = col(df, "Elevate Status", "Elevate Customers", "Elevate purchased")
    C_RESOLVE = col(df, "Resolve Status", "Resolve purchased")
    C_CRM     = col(df, "CRM Importer", "SOR")
    C_SSO     = col(df, "SSO Provider")
    C_WB      = col(df, "Write backs")
    C_NOTES   = col(df, "Notes")
    C_URL     = col(df, "Customer URLs")

    # Clean
    if C_CUST:
        df = df[df[C_CUST].notna() & (df[C_CUST].str.strip() != "")]

    # ── Filters row ───────────────────────────────────────────────────────────
    fc1, fc2, fc3, fc4 = st.columns([1,1,1,1])
    tse_opts = ["All"] + sorted(df[C_TSE].dropna().unique().tolist()) if C_TSE else ["All"]
    csm_opts = ["All"] + sorted(df[C_CSM].dropna().unique().tolist()) if C_CSM else ["All"]
    stage_opts = ["All"] + sorted(df[C_STAGE].dropna().unique().tolist()) if C_STAGE else ["All"]

    sel_tse   = fc1.selectbox("TSE",   tse_opts,   key="tam_tse2")
    sel_csm   = fc2.selectbox("CSM",   csm_opts,   key="tam_csm2")
    sel_stage = fc3.selectbox("Stage", stage_opts, key="tam_stage")
    search    = fc4.text_input("Search account", placeholder="type name...", key="tam_search")

    dff = df.copy()
    if sel_tse   != "All" and C_TSE:   dff = dff[dff[C_TSE].str.strip()==sel_tse]
    if sel_csm   != "All" and C_CSM:   dff = dff[dff[C_CSM].str.strip()==sel_csm]
    if sel_stage != "All" and C_STAGE: dff = dff[dff[C_STAGE].str.strip()==sel_stage]
    if search and C_CUST:
        dff = dff[dff[C_CUST].str.lower().str.contains(search.lower(), na=False)]

    # ── KPI row ───────────────────────────────────────────────────────────────
    total = len(dff)
    arr_sum = pd.to_numeric(dff[C_ARR], errors="coerce").sum() if C_ARR else 0
    live_n  = dff[C_STAGE].str.lower().str.contains("adopt|live", na=False).sum() if C_STAGE else 0

    today = pd.Timestamp.now()
    renew_soon = 0
    if C_RENEW:
        def parse_dt(v):
            try: return pd.to_datetime(v)
            except: return None
        rens = dff[C_RENEW].apply(parse_dt)
        renew_soon = ((rens >= today) & (rens <= today + pd.Timedelta(days=90))).sum()

    k1,k2,k3,k4 = st.columns(4)
    k1.metric("Accounts", total)
    k2.metric("Total ARR", f"${arr_sum/1000:.0f}K" if arr_sum else "—")
    k3.metric("Live / Adoption", live_n)
    k4.metric("Renewals in 90d", int(renew_soon), "⚠️" if renew_soon else "")

    st.divider()

    # ── Interactive account + module matrix ───────────────────────────────────
    st.subheader("📋 Account & module status")
    st.caption("Each row = one account. Module columns show live status. "
               "Click column headers to sort. Use filters above to narrow down.")

    def status_icon(v):
        if not v or str(v).strip() in ("","--","nan","None"): return "⬜"
        s = str(v).lower()
        if "churn"   in s: return "❌"
        if "live"    in s: return "✅"
        if "impl"    in s: return "🔧"
        if "entitl"  in s: return "📋"
        if "premier" in s: return "✅"
        if "manual"  in s: return "🔧"
        return "⚪"

    def days_to_renewal(v):
        try:
            dt = pd.to_datetime(v)
            d = (dt - today).days
            if d < 0:   return f"🔴 {abs(d)}d overdue"
            if d <= 90: return f"🟡 {d}d"
            return f"🟢 {d}d"
        except:
            return str(v)[:10] if v else "—"

    # Build display dataframe
    rows_out = []
    for _, row in dff.iterrows():
        cust  = str(row.get(C_CUST,"")).strip()  if C_CUST  else "—"
        tse   = str(row.get(C_TSE,"")).strip()   if C_TSE   else "—"
        csm   = str(row.get(C_CSM,"")).strip()   if C_CSM   else "—"
        stage = str(row.get(C_STAGE,"")).strip() if C_STAGE else "—"
        arr   = row.get(C_ARR,"")
        renew = row.get(C_RENEW,"")
        try:   arr_fmt = f"${float(arr)/1000:.0f}K"
        except: arr_fmt = "—"
        rows_out.append({
            "Account":   cust,
            "TSE":       tse,
            "CSM":       csm,
            "Stage":     stage,
            "ARR":       arr_fmt,
            "Renewal":   days_to_renewal(renew),
            "Core":      status_icon(row.get(C_CORE))    if C_CORE    else "—",
            "Assign":    status_icon(row.get(C_ASSIGN))  if C_ASSIGN  else "—",
            "Assist":    status_icon(row.get(C_ASSIST))  if C_ASSIST  else "—",
            "Expand":    status_icon(row.get(C_EXPAND))  if C_EXPAND  else "—",
            "Elevate":   status_icon(row.get(C_ELEVATE)) if C_ELEVATE else "—",
            "Resolve":   status_icon(row.get(C_RESOLVE)) if C_RESOLVE else "—",
        })

    df_display = pd.DataFrame(rows_out)
    st.dataframe(df_display, use_container_width=True, hide_index=True,
                 height=min(60 + len(df_display)*35, 600))

    # Legend
    st.markdown(
        "<div style='font-size:10px;color:#8b949e;margin-top:4px'>"
        "✅ Live &nbsp;&nbsp; 🔧 Implementing &nbsp;&nbsp; 📋 Entitled (not live) &nbsp;&nbsp; "
        "⬜ Not purchased &nbsp;&nbsp; ❌ Churned &nbsp;&nbsp; ⚪ Other"
        "</div>", unsafe_allow_html=True)

    st.divider()

    # ── Account detail expander ───────────────────────────────────────────────
    st.subheader("🔍 Account detail")
    acct_list = dff[C_CUST].dropna().tolist() if C_CUST else []
    sel_acct  = st.selectbox("Select account for full detail", ["—"] + acct_list,
                              key="tam_acct_detail")

    if sel_acct != "—" and C_CUST:
        row = dff[dff[C_CUST]==sel_acct].iloc[0]
        notes = str(row.get(C_NOTES,"")).strip() if C_NOTES else ""
        url   = str(row.get(C_URL,"")).strip()   if C_URL   else ""
        crm   = str(row.get(C_CRM,"")).strip()   if C_CRM   else "—"
        sso   = str(row.get(C_SSO,"")).strip()   if C_SSO   else "—"
        wb    = str(row.get(C_WB,"")).strip()    if C_WB    else "—"

        col_d1, col_d2 = st.columns(2)
        with col_d1:
            st.markdown(f"""
<div style='background:#161b22;border-radius:10px;padding:16px 20px;border:0.5px solid #30363d'>
  <div style='font-size:16px;font-weight:700;color:#e6edf3;margin-bottom:12px'>{sel_acct}</div>
  <div style='display:grid;grid-template-columns:1fr 1fr;gap:8px'>
    <div><div style='font-size:10px;color:#8b949e'>TSE</div><div style='color:#e6edf3;font-size:12px'>{str(row.get(C_TSE,"")).strip() if C_TSE else "—"}</div></div>
    <div><div style='font-size:10px;color:#8b949e'>CSM</div><div style='color:#e6edf3;font-size:12px'>{str(row.get(C_CSM,"")).strip() if C_CSM else "—"}</div></div>
    <div><div style='font-size:10px;color:#8b949e'>ARR</div><div style='color:#1D9E75;font-size:13px;font-weight:600'>{try_arr(row.get(C_ARR)) if C_ARR else "—"}</div></div>
    <div><div style='font-size:10px;color:#8b949e'>Stage</div><div style='color:#e6edf3;font-size:12px'>{str(row.get(C_STAGE,"")).strip() if C_STAGE else "—"}</div></div>
    <div><div style='font-size:10px;color:#8b949e'>Renewal</div><div style='color:#e6edf3;font-size:12px'>{days_to_renewal(row.get(C_RENEW)) if C_RENEW else "—"}</div></div>
    <div><div style='font-size:10px;color:#8b949e'>CRM</div><div style='color:#e6edf3;font-size:12px'>{crm}</div></div>
    <div><div style='font-size:10px;color:#8b949e'>SSO</div><div style='color:#e6edf3;font-size:12px'>{sso}</div></div>
    <div><div style='font-size:10px;color:#8b949e'>Writeback</div><div style='color:#e6edf3;font-size:12px'>{wb}</div></div>
  </div>
  {"<div style=\'margin-top:10px\'><a href=\'"+url+"\' target=\'_blank\' style=\'color:#185FA5;font-size:11px\'>🔗 Open SL portal</a></div>" if url and url not in ("nan","None") else ""}
</div>""", unsafe_allow_html=True)

        with col_d2:
            st.markdown("**Module status**")
            for mod_name, mod_col_key in [
                ("Core SX",   C_CORE), ("Assign",    C_ASSIGN),
                ("Assist",    C_ASSIST), ("Expand",  C_EXPAND),
                ("Elevate",   C_ELEVATE), ("Resolve",C_RESOLVE)
            ]:
                val = str(row.get(mod_col_key,"")).strip() if mod_col_key else "—"
                ic  = status_icon(val)
                val_disp = val if val not in ("nan","None","--","") else "—"
                s = val.lower()
                if "live" in s or "premier" in s: bc="#1D9E75"; bg="#0f2a1a"
                elif "impl" in s: bc="#BA7517"; bg="#2a2a10"
                elif "entitl" in s: bc="#534AB7"; bg="#1a1a3a"
                elif "churn" in s: bc="#E24B4A"; bg="#3a1a1a"
                else: bc="#444"; bg="#1e1e1e"
                st.markdown(
                    f"<div style='background:{bg};border:0.5px solid {bc};"
                    f"border-radius:6px;padding:6px 10px;margin-bottom:4px;"
                    f"display:flex;justify-content:space-between'>"
                    f"<span style='font-size:11px;color:#e6edf3'>{mod_name}</span>"
                    f"<span style='font-size:11px;color:{bc}'>{ic} {val_disp}</span></div>",
                    unsafe_allow_html=True)

        if notes and notes not in ("nan","None",""):
            st.markdown("**Notes**")
            st.markdown(
                f"<div style='background:#161b22;border-radius:8px;padding:12px 16px;"
                f"font-size:12px;color:#8b949e;line-height:1.6;border:0.5px solid #30363d'>"
                f"{notes[:500] + '…' if len(notes)>500 else notes}</div>",
                unsafe_allow_html=True)

    st.divider()

    # ── Ongoing projects ──────────────────────────────────────────────────────
    st.subheader("🔧 Ongoing projects")
    df_proj = proj_data.get("Ongoing Projects", pd.DataFrame())
    if not df_proj.empty:
        PC_OWN  = col(df_proj,"Project Owner")
        PC_CUST = col(df_proj,"Customer")
        PC_PROG = col(df_proj,"Progress")
        PC_TASK = col(df_proj,"Task/Activity","Task")
        PC_STAT = col(df_proj,"Current Status")
        PC_NEXT = col(df_proj,"Next Steps")

        df_proj_f = df_proj.copy()
        if sel_tse != "All" and PC_OWN:
            df_proj_f = df_proj_f[df_proj_f[PC_OWN].str.lower().str.contains(
                sel_tse.lower().split()[0], na=False)]

        STATUS_C = {"in progress":"#1D9E75","not started":"#BA7517",
                    "scoping":"#185FA5","done":"#534AB7","blocked":"#E24B4A"}

        for _, pr in df_proj_f.iterrows():
            task  = str(pr.get(PC_TASK,"")).strip() if PC_TASK else "—"
            pcust = str(pr.get(PC_CUST,"")).strip() if PC_CUST else "—"
            prog  = str(pr.get(PC_PROG,"")).strip() if PC_PROG else "—"
            stat  = str(pr.get(PC_STAT,"")).strip() if PC_STAT else ""
            nxt   = str(pr.get(PC_NEXT,"")).strip() if PC_NEXT else ""
            own   = str(pr.get(PC_OWN,"")).strip()  if PC_OWN  else ""
            if task in ("nan","None","") and pcust in ("nan","None",""): continue
            pc = STATUS_C.get(prog.lower(),"#8b949e")
            st.markdown(
                f"<div style='background:#161b22;border:0.5px solid #30363d;"
                f"border-left:3px solid {pc};border-radius:8px;"
                f"padding:10px 16px;margin-bottom:6px'>"
                f"<div style='display:flex;justify-content:space-between'>"
                f"<span style='color:#e6edf3;font-size:13px;font-weight:500'>"
                f"{task if task not in ('nan','None') else '—'}</span>"
                f"<span style='color:#8b949e;font-size:10px'>{pcust} · {own}</span></div>"
                f"<span style='color:{pc};font-size:10px;font-weight:600'>{prog}</span>"
                + (f"<div style='color:#8b949e;font-size:11px;margin-top:4px'>📌 {stat}</div>" if stat and stat not in ("nan","None") else "")
                + (f"<div style='color:#185FA5;font-size:11px;margin-top:2px'>→ {nxt}</div>" if nxt and nxt not in ("nan","None") else "")
                + "</div>", unsafe_allow_html=True)
    else:
        st.info("No ongoing projects data.", icon="ℹ️")

    st.caption(f"Google Sheet ID: {SHEET_ID} · Cached 5 min · {len(dff)} accounts shown")




elif tab == 10:
    import openpyxl
    import pandas as pd
    from pathlib import Path
    from utils.charts import COLORS, bar_chart, donut_chart

    st.markdown("""
<div style='display:flex;align-items:center;gap:12px;margin-bottom:4px'>
  <div style='font-size:20px;font-weight:700;color:#e6edf3'>👤 TAM / TSE Account Dashboard</div>
  <div style='font-size:11px;color:#8b949e;margin-top:2px'>
    Source: accounts.xlsx · Select a TSE to view their portfolio
  </div>
</div>""", unsafe_allow_html=True)

    # ── Load accounts.xlsx ────────────────────────────────────────────────────
    ACCOUNTS_FILE = Path("data/accounts.xlsx")
    if not ACCOUNTS_FILE.exists():
        st.error(
            "accounts.xlsx not found at `data/accounts.xlsx`. "
            "Copy the file from Downloads:\n"
            "```bash\nmkdir -p data && cp ~/Downloads/accounts.xlsx data/\n```"
        )
        st.stop()

    @st.cache_data(ttl=3600, show_spinner=False)
    def load_accounts():
        wb  = openpyxl.load_workbook("data/accounts.xlsx", read_only=True, data_only=True)

        # ── Up to date sheet ──────────────────────────────────────────────
        ws  = wb["Up to date"]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip() if h else f"col_{i}" for i,h in enumerate(rows[0])]
        data = []
        for row in rows[1:]:
            if any(v is not None for v in row):
                data.append(dict(zip(headers, row)))
        df_main = pd.DataFrame(data)

        # ── Account Contact List ──────────────────────────────────────────
        ws2 = wb["Account Contact List"]
        rows2 = list(ws2.iter_rows(values_only=True))
        h2 = [str(h).strip() if h else f"col_{i}" for i,h in enumerate(rows2[0])]
        d2 = [dict(zip(h2, r)) for r in rows2[1:] if any(v is not None for v in r)]
        df_contacts = pd.DataFrame(d2)

        # ── Ongoing Projects ──────────────────────────────────────────────
        ws3 = wb["Ongoing Projects"]
        rows3 = list(ws3.iter_rows(values_only=True))
        h3 = [str(h).strip() if h else f"col_{i}" for i,h in enumerate(rows3[1])]
        d3 = [dict(zip(h3, r)) for r in rows3[2:] if any(v is not None for v in r)]
        df_projects = pd.DataFrame(d3)

        # ── Incoming ──────────────────────────────────────────────────────
        ws4 = wb["Incoming"]
        rows4 = list(ws4.iter_rows(values_only=True))
        h4 = [str(h).strip() if h else f"col_{i}" for i,h in enumerate(rows4[1])]
        d4 = [dict(zip(h4, r)) for r in rows4[2:] if any(v is not None for v in r)]
        df_incoming = pd.DataFrame(d4)

        wb.close()
        return df_main, df_contacts, df_projects, df_incoming

    with st.spinner("Loading account data..."):
        df_main, df_contacts, df_projects, df_incoming = load_accounts()

    # ── Normalize key columns ─────────────────────────────────────────────────
    def col(df, *candidates):
        """Find first matching column name."""
        for c in candidates:
            if c in df.columns:
                return c
            matches = [x for x in df.columns if c.lower() in str(x).lower()]
            if matches:
                return matches[0]
        return None

    tse_col     = col(df_main, "Assigned TSE", "TSE")
    csm_col     = col(df_main, "Customer Success Manager", "CSM")
    cust_col    = col(df_main, "Customer name", "Customer")
    arr_col     = col(df_main, "ARR")
    renewal_col = col(df_main, "Renewal Date")
    stage_col   = col(df_main, "Lifecycle Stage")
    core_col    = col(df_main, "Core Status")
    assign_col  = col(df_main, "Assign Status", "Assign purchased")
    assist_col  = col(df_main, "Assist Status", "Assist purchased")
    expand_col  = col(df_main, "Expand Status", "Expand purchased")
    elevate_col = col(df_main, "Elevate Status", "Elevate Customers", "Elevate purchased")
    resolve_col = col(df_main, "Resolve Status", "Resolve purchased")
    crm_col     = col(df_main, "CRM Importer", "SOR")
    sso_col     = col(df_main, "SSO Provider")
    wb_col      = col(df_main, "Write backs")
    notes_col   = col(df_main, "Notes")
    url_col     = col(df_main, "Customer URLs")

    # ── TSE selector ──────────────────────────────────────────────────────────
    all_tses = sorted(df_main[tse_col].dropna().unique().tolist()) if tse_col else []
    col_tse, col_csm, _ = st.columns([1, 1, 2])
    selected_tse = col_tse.selectbox("Select TSE", ["All"] + all_tses, key="tam_tse")
    all_csms = sorted(df_main[csm_col].dropna().unique().tolist()) if csm_col else []
    selected_csm = col_csm.selectbox("Filter by CSM", ["All"] + all_csms, key="tam_csm")

    # Filter
    df_filtered = df_main.copy()
    if selected_tse != "All" and tse_col:
        df_filtered = df_filtered[df_filtered[tse_col].astype(str).str.strip() == selected_tse]
    if selected_csm != "All" and csm_col:
        df_filtered = df_filtered[df_filtered[csm_col].astype(str).str.strip() == selected_csm]
    df_filtered = df_filtered[df_filtered[cust_col].notna()] if cust_col else df_filtered

    st.divider()

    # ── Portfolio KPIs ────────────────────────────────────────────────────────
    total_accounts = len(df_filtered)
    total_arr = 0
    if arr_col:
        arr_vals = pd.to_numeric(df_filtered[arr_col], errors='coerce').dropna()
        total_arr = arr_vals.sum()

    live_count = 0
    if stage_col:
        live_count = df_filtered[stage_col].astype(str).str.lower().str.contains("adopt|live|impl").sum()

    # Renewals in next 90 days
    renewal_soon = 0
    if renewal_col:
        today = pd.Timestamp.now()
        def parse_renewal(v):
            if isinstance(v, datetime_type := __import__('datetime').datetime):
                return pd.Timestamp(v)
            try:
                return pd.to_datetime(v)
            except:
                return None
        import datetime
        renewals = df_filtered[renewal_col].apply(parse_renewal)
        renewal_soon = ((renewals >= today) & (renewals <= today + pd.Timedelta(days=90))).sum()

    c1,c2,c3,c4,c5 = st.columns(5)
    c1.metric("Total accounts", total_accounts)
    c2.metric("Total ARR", f"${total_arr/1000:.0f}K" if total_arr else "—")
    c3.metric("Live / Adoption", live_count)
    c4.metric("Renewals in 90d", renewal_soon, "⚠️" if renewal_soon > 0 else "")
    # Projects for this TSE
    proj_count = 0
    if "Project Owner" in df_projects.columns and selected_tse != "All":
        proj_count = df_projects[df_projects["Project Owner"].astype(str).str.lower().str.contains(
            selected_tse.lower().split()[0] if selected_tse else "", na=False)].shape[0]
    c5.metric("Active projects", proj_count)

    st.divider()

    # ── Account cards ─────────────────────────────────────────────────────────
    st.subheader(f"📋 Account portfolio — {selected_tse if selected_tse != 'All' else 'All TSEs'}")

    STATUS_ICON = {
        "live":         ("✅", "#1a3a2a", "#1D9E75"),
        "implementing": ("🔧", "#2a2a10", "#BA7517"),
        "entitled":     ("📋", "#1a1a3a", "#534AB7"),
        "not":          ("⬜", "#1e1e1e", "#444"),
        "churned":      ("❌", "#3a1a1a", "#E24B4A"),
        "scoping":      ("🔍", "#1a1a3a", "#185FA5"),
        "other":        ("⚪", "#161b22", "#30363d"),
    }

    def status_class(val):
        if not val or str(val).strip() in ("--","None","nan",""):
            return "not"
        v = str(val).lower()
        if "churn" in v:  return "churned"
        if "live" in v:   return "live"
        if "impl" in v:   return "implementing"
        if "entitl" in v: return "entitled"
        if "scop" in v:   return "scoping"
        return "other"

    def module_badge(label, val):
        sc = status_class(val)
        icon, bg, border = STATUS_ICON[sc]
        disp = str(val).replace("--","—") if val and str(val) not in ("nan","None","--") else "—"
        return (f"<div style='background:{bg};border:0.5px solid {border};"
                f"border-radius:6px;padding:4px 8px;font-size:10px;margin:2px'>"
                f"<span style='color:{border}'>{icon}</span> "
                f"<span style='color:#8b949e'>{label}</span> "
                f"<span style='color:#e6edf3;font-weight:500'>{disp}</span></div>")

    for _, row in df_filtered.iterrows():
        cust_name = str(row.get(cust_col,"")).strip() if cust_col else "—"
        if not cust_name or cust_name in ("nan","None"):
            continue

        stage  = str(row.get(stage_col,"")).strip()  if stage_col  else ""
        arr    = row.get(arr_col)                     if arr_col    else None
        renew  = row.get(renewal_col)                 if renewal_col else None
        tse    = str(row.get(tse_col,"")).strip()     if tse_col    else "—"
        csm    = str(row.get(csm_col,"")).strip()     if csm_col    else "—"
        notes  = str(row.get(notes_col,"")).strip()   if notes_col  else ""
        url    = str(row.get(url_col,"")).strip()     if url_col    else ""
        crm    = str(row.get(crm_col,"")).strip()     if crm_col    else "—"
        sso    = str(row.get(sso_col,"")).strip()     if sso_col    else "—"
        wb_val = str(row.get(wb_col,"")).strip()      if wb_col     else "—"

        # Renewal date formatting
        renew_str = "—"
        renew_flag = ""
        if renew and str(renew) not in ("nan","None",""):
            try:
                import datetime
                if isinstance(renew, datetime.datetime):
                    rdt = pd.Timestamp(renew)
                else:
                    rdt = pd.to_datetime(renew)
                renew_str = rdt.strftime("%b %Y")
                days_to = (rdt - pd.Timestamp.now()).days
                if days_to < 0:
                    renew_flag = "🔴 Overdue"
                elif days_to <= 90:
                    renew_flag = f"🟡 {days_to}d away"
                else:
                    renew_flag = f"🟢 {days_to}d away"
            except:
                renew_str = str(renew)[:10]

        # ARR formatting
        arr_str = f"${float(arr)/1000:.0f}K" if arr and str(arr) not in ("nan","None") else "—"

        # Stage color
        stage_lower = stage.lower()
        if "adopt" in stage_lower or "live" in stage_lower:
            stage_color = "#1D9E75"
        elif "impl" in stage_lower:
            stage_color = "#BA7517"
        elif "churn" in stage_lower:
            stage_color = "#E24B4A"
        else:
            stage_color = "#534AB7"

        # Module statuses
        modules = {
            "Core":    row.get(core_col)    if core_col    else None,
            "Assign":  row.get(assign_col)  if assign_col  else None,
            "Assist":  row.get(assist_col)  if assist_col  else None,
            "Expand":  row.get(expand_col)  if expand_col  else None,
            "Elevate": row.get(elevate_col) if elevate_col else None,
            "Resolve": row.get(resolve_col) if resolve_col else None,
        }

        # Notes truncation
        notes_disp = (notes[:180] + "…") if len(notes) > 180 else notes
        notes_disp = notes_disp if notes_disp not in ("nan","None","") else ""

        url_link = f"<a href='{url}' target='_blank' style='color:#185FA5;font-size:10px'>🔗 Open SL portal</a>" if url and url not in ("nan","None") else ""

        badges = "".join(module_badge(k,v) for k,v in modules.items())

        with st.expander(
            f"**{cust_name}**  ·  "
            f"<span style='color:{stage_color}'>{stage}</span>  ·  "
            f"{arr_str}  ·  Renewal: {renew_str} {renew_flag}  ·  TSE: {tse}",
            expanded=False
        ):
            col_l, col_r = st.columns([3, 2])
            with col_l:
                st.markdown(f"""
<div style='background:#161b22;border-radius:8px;padding:12px 16px;margin-bottom:8px'>
  <div style='display:flex;gap:24px;flex-wrap:wrap;margin-bottom:8px'>
    <div><span style='font-size:10px;color:#8b949e'>CSM</span><br><span style='color:#e6edf3;font-size:12px'>{csm}</span></div>
    <div><span style='font-size:10px;color:#8b949e'>TSE</span><br><span style='color:#e6edf3;font-size:12px'>{tse}</span></div>
    <div><span style='font-size:10px;color:#8b949e'>ARR</span><br><span style='color:#1D9E75;font-size:12px;font-weight:600'>{arr_str}</span></div>
    <div><span style='font-size:10px;color:#8b949e'>Renewal</span><br><span style='color:#e6edf3;font-size:12px'>{renew_str}</span> <span style='font-size:10px'>{renew_flag}</span></div>
    <div><span style='font-size:10px;color:#8b949e'>CRM</span><br><span style='color:#e6edf3;font-size:12px'>{crm}</span></div>
    <div><span style='font-size:10px;color:#8b949e'>SSO</span><br><span style='color:#e6edf3;font-size:12px'>{sso}</span></div>
    <div><span style='font-size:10px;color:#8b949e'>Writeback</span><br><span style='color:#e6edf3;font-size:12px'>{wb_val}</span></div>
  </div>
  {url_link}
  {"<div style='margin-top:8px;font-size:11px;color:#8b949e;line-height:1.5'>" + notes_disp + "</div>" if notes_disp else ""}
</div>""", unsafe_allow_html=True)

            with col_r:
                st.markdown("**Module status**")
                st.markdown(f"<div style='display:flex;flex-wrap:wrap;gap:2px'>{badges}</div>",
                            unsafe_allow_html=True)

    st.divider()

    # ── Ongoing Projects ──────────────────────────────────────────────────────
    st.subheader("🔧 Ongoing projects")
    if not df_projects.empty:
        proj_col_owner = col(df_projects, "Project Owner")
        proj_col_cust  = col(df_projects, "Customer")
        proj_col_prog  = col(df_projects, "Progress")
        proj_col_task  = col(df_projects, "Task/Activity", "Task")
        proj_col_status= col(df_projects, "Current Status")
        proj_col_next  = col(df_projects, "Next Steps")
        proj_col_start = col(df_projects, "Start Date")
        proj_col_age   = col(df_projects, "Project Age")

        df_proj_show = df_projects.copy()
        if selected_tse != "All" and proj_col_owner:
            df_proj_show = df_proj_show[
                df_proj_show[proj_col_owner].astype(str).str.lower().str.contains(
                    selected_tse.lower().split()[0], na=False)]

        STATUS_COLORS = {
            "in progress": "#1D9E75",
            "not started": "#BA7517",
            "scoping":     "#185FA5",
            "done":        "#534AB7",
            "blocked":     "#E24B4A",
        }

        for _, proj in df_proj_show.iterrows():
            owner  = str(proj.get(proj_col_owner,"")).strip()  if proj_col_owner  else "—"
            pcust  = str(proj.get(proj_col_cust,"")).strip()   if proj_col_cust   else "—"
            prog   = str(proj.get(proj_col_prog,"")).strip()   if proj_col_prog   else "—"
            task   = str(proj.get(proj_col_task,"")).strip()   if proj_col_task   else "—"
            status = str(proj.get(proj_col_status,"")).strip() if proj_col_status else ""
            nexts  = str(proj.get(proj_col_next,"")).strip()   if proj_col_next   else ""

            if task in ("nan","None","") and pcust in ("nan","None",""):
                continue

            prog_c = STATUS_COLORS.get(prog.lower(), "#8b949e")
            st.markdown(f"""
<div style='background:#161b22;border:0.5px solid #30363d;border-radius:8px;
     padding:12px 16px;margin-bottom:6px;border-left:3px solid {prog_c}'>
  <div style='display:flex;justify-content:space-between;align-items:flex-start'>
    <div>
      <span style='font-size:13px;font-weight:500;color:#e6edf3'>{task if task not in ("nan","None") else "—"}</span>
      <span style='font-size:11px;color:#8b949e;margin-left:10px'>{pcust if pcust not in ("nan","None") else ""}</span>
    </div>
    <div style='display:flex;gap:8px;align-items:center'>
      <span style='font-size:10px;color:{prog_c};font-weight:600'>{prog}</span>
      <span style='font-size:10px;color:#8b949e'>{owner}</span>
    </div>
  </div>
  {"<div style='font-size:11px;color:#8b949e;margin-top:6px'>📌 " + status + "</div>" if status and status not in ("nan","None") else ""}
  {"<div style='font-size:11px;color:#185FA5;margin-top:4px'>→ " + nexts + "</div>" if nexts and nexts not in ("nan","None") else ""}
</div>""", unsafe_allow_html=True)
    else:
        st.info("No ongoing projects data available.", icon="ℹ️")

    st.divider()

    # ── Incoming projects ─────────────────────────────────────────────────────
    st.subheader("📥 Incoming / pipeline")
    if not df_incoming.empty:
        inc_cust = col(df_incoming, "Customer")
        inc_task = col(df_incoming, "Task/Activity", "Task")
        inc_prog = col(df_incoming, "Progress")
        inc_pri  = col(df_incoming, "Priority")
        inc_own  = col(df_incoming, "Potential Project Owner")

        for _, inc in df_incoming.iterrows():
            icust = str(inc.get(inc_cust,"")).strip() if inc_cust else "—"
            itask = str(inc.get(inc_task,"")).strip() if inc_task else "—"
            iprog = str(inc.get(inc_prog,"")).strip() if inc_prog else "—"
            ipri  = str(inc.get(inc_pri,"")).strip()  if inc_pri  else "—"
            iown  = str(inc.get(inc_own,"")).strip()  if inc_own  else "—"
            if icust in ("nan","None","") and itask in ("nan","None",""):
                continue
            pri_c = "#E24B4A" if ipri.lower()=="high" else "#BA7517" if ipri.lower()=="medium" else "#8b949e"
            st.markdown(f"""
<div style='background:#161b22;border:0.5px solid #30363d;border-radius:8px;
     padding:10px 16px;margin-bottom:6px'>
  <div style='display:flex;justify-content:space-between'>
    <span style='font-size:12px;font-weight:500;color:#e6edf3'>
      {icust if icust not in ("nan","None") else "—"} — {itask if itask not in ("nan","None") else "—"}
    </span>
    <span style='font-size:10px;color:{pri_c};font-weight:600'>{ipri} priority</span>
  </div>
  <div style='font-size:10px;color:#8b949e;margin-top:4px'>{iprog} · Owner: {iown}</div>
</div>""", unsafe_allow_html=True)
    else:
        st.info("No incoming projects data.", icon="ℹ️")

    st.divider()

    # ── Portfolio summary charts ──────────────────────────────────────────────
    st.subheader("📊 Portfolio overview")
    col_ch1, col_ch2, col_ch3 = st.columns(3)

    # Stage breakdown
    if stage_col and not df_filtered.empty:
        stage_counts = df_filtered[stage_col].value_counts().reset_index()
        stage_counts.columns = ["stage","count"]
        stage_counts = stage_counts[stage_counts["stage"].notna()]
        with col_ch1:
            st.markdown("**Lifecycle stage**")
            st.plotly_chart(donut_chart(
                labels=stage_counts["stage"].tolist(),
                values=stage_counts["count"].tolist(),
                colors=[COLORS["teal"],COLORS["blue"],COLORS["amber"],
                        COLORS["red"],COLORS["gray"],COLORS["purple"]],
                height=220,
                center_text=f"{len(df_filtered)} accounts"
            ), use_container_width=True)

    # Module adoption
    module_live = {}
    for mod_name, mod_col in [
        ("Core", core_col), ("Assign", assign_col), ("Assist", assist_col),
        ("Expand", expand_col), ("Elevate", elevate_col), ("Resolve", resolve_col)
    ]:
        if mod_col and mod_col in df_filtered.columns:
            live = df_filtered[mod_col].astype(str).str.lower().str.contains("live|active").sum()
            module_live[mod_name] = int(live)

    if module_live:
        mod_df = pd.DataFrame(list(module_live.items()), columns=["module","live_count"])
        with col_ch2:
            st.markdown("**Modules live across portfolio**")
            st.plotly_chart(bar_chart(
                mod_df, x="module",
                y_cols=[{"col":"live_count","name":"Live accounts",
                          "color":[COLORS["teal"],COLORS["blue"],COLORS["purple"],
                                   COLORS["amber"],COLORS["green"],COLORS["red"]],
                          "colors":[COLORS["teal"],COLORS["blue"],COLORS["purple"],
                                    COLORS["amber"],COLORS["green"],COLORS["red"]]}],
                height=220
            ), use_container_width=True)

    # ARR by account (top 10)
    if arr_col and not df_filtered.empty:
        arr_df = df_filtered[[cust_col, arr_col]].copy()
        arr_df.columns = ["customer","arr"]
        arr_df["arr"] = pd.to_numeric(arr_df["arr"], errors="coerce")
        arr_df = arr_df.dropna().sort_values("arr", ascending=False).head(10)
        arr_df["arr_k"] = (arr_df["arr"]/1000).round(0)
        arr_df["customer"] = arr_df["customer"].str[:20]
        with col_ch3:
            st.markdown("**ARR by account (top 10, $K)**")
            st.plotly_chart(bar_chart(
                arr_df.sort_values("arr_k"), x="customer",
                y_cols=[{"col":"arr_k","name":"ARR ($K)","color":COLORS["teal"]}],
                height=220, horizontal=True
            ), use_container_width=True)

    st.caption(f"Source: data/accounts.xlsx · {len(df_filtered)} accounts shown · "
               f"TSE filter: {selected_tse} · CSM filter: {selected_csm}")

